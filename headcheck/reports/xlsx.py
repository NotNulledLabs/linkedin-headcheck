"""
Excel (.xlsx) workbook tuned for HR review.

Colour-coded risk cells, score data bars, hyperlinked profile URLs,
frozen header, autofilter, empty Notes column for inline annotation,
and a hidden _meta sheet with provenance info.
"""
from datetime import datetime

from ..constants import VERSION, MAX_SCORE


_XLSX_RISK_FILL = {
    "red":    "FCE4E4",   # light red
    "yellow": "FFF4C2",   # light amber
    "green":  "DCEEDC",   # light green
}

# Risk labels are kept human-readable in the sheet. The raw code is available
# elsewhere via the 'risk' column, but labels are what HR will scan.
_XLSX_RISK_LABEL = {
    "red":    "🔴 High risk",
    "yellow": "🟡 Needs review",
    "green":  "🟢 Low risk",
}


def generate_xlsx(profiles: list[dict], company: str, has_payroll: bool,
                  out: str) -> int:
    """
    Write the full profile list to an Excel workbook tuned for HR review:
    frozen header row, autofilter, colour-coded risk column, score data
    bars, hyperlinked profile URLs, and an empty Notes column for inline
    annotation. Returns the number of profile rows written.

    This report is complementary to the suspects CSV (red+yellow only) and
    to the interactive HTML (all profiles, with search/filter in-browser).
    The XLSX exists for HR teams who work in Excel end-to-end: they can
    sort, filter, and annotate without leaving their usual tool.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HeadCheck"

    # ── Header row ──
    headers = [
        "Risk", "Score", "Name", "Headline", "Profile URL",
        "Photo state", "Mutual level", "Generic slug?",
        "Suspicious name?", "Suspicious reason",
    ]
    if has_payroll:
        headers += ["Payroll status", "Payroll match"]
    headers += ["HR notes"]

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A2F3A")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DE"),
        right=Side(style="thin", color="D5D8DE"),
        top=Side(style="thin", color="D5D8DE"),
        bottom=Side(style="thin", color="D5D8DE"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = thin_border

    # ── Data rows, ordered worst-first so the interesting ones appear on top ──
    # Worst-first ordering. Defensive against name being None or missing —
    # we use .get() with a string default so .lower() doesn't raise.
    ordered = sorted(profiles, key=lambda p: (p["score"], (p.get("name") or "").lower()))

    for p in ordered:
        row = [
            _XLSX_RISK_LABEL.get(p["risk"], p["risk"]),
            p["score"],
            p["name"],
            p.get("headline", ""),
            p.get("profile_url", ""),
            p.get("photo_state", "loaded" if p.get("has_photo") else "absent"),
            p.get("mutual_level", 0),
            "yes" if p.get("slug_is_generic") else "no",
            "yes" if p.get("suspicious_name") else "no",
            p.get("suspicious_reason", ""),
        ]
        if has_payroll:
            row += [p.get("payroll_status", ""), p.get("payroll_match", "")]
        row += [""]   # empty Notes column for HR to fill in

        ws.append(row)
        r = ws.max_row

        # Risk cell background — makes the sheet scannable at a glance.
        ws.cell(row=r, column=1).fill = PatternFill(
            "solid", fgColor=_XLSX_RISK_FILL.get(p["risk"], "FFFFFF")
        )

        # Profile URL as hyperlink — HR can click through from the sheet.
        url_cell = ws.cell(row=r, column=5)
        if p.get("profile_url"):
            url_cell.hyperlink = p["profile_url"]
            url_cell.font = Font(color="2563EB", underline="single")

    # ── Visuals: data bars on the score column ──
    # Score column is #2; data rows start at row 2. Use row count for range.
    last_row = ws.max_row
    if last_row >= 2:
        score_range = f"B2:B{last_row}"
        ws.conditional_formatting.add(
            score_range,
            DataBarRule(start_type="num", start_value=0,
                        end_type="num", end_value=MAX_SCORE,
                        color="4CAF50", showValue=True),
        )

    # ── Freeze header + enable autofilter so HR can sort/filter ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Column widths: set once based on header length + generous padding.
    # Fine-tuning is manual work HR can do from Excel itself if needed.
    widths = {
        "A": 18,   # Risk
        "B": 8,    # Score
        "C": 28,   # Name
        "D": 50,   # Headline
        "E": 45,   # URL
        "F": 13,   # Photo state
        "G": 14,   # Mutual level
        "H": 14,   # Generic slug?
        "I": 17,   # Suspicious name?
        "J": 28,   # Suspicious reason
    }
    if has_payroll:
        widths.update({"K": 16, "L": 28})                                 # payroll cols
        widths[get_column_letter(13)] = 40                                # notes
    else:
        widths[get_column_letter(11)] = 40                                # notes
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # A tiny metadata sheet — for future diffs and provenance.
    meta = wb.create_sheet("_meta")
    meta["A1"] = "Generated by"
    meta["B1"] = f"LinkedIn HeadCheck v{VERSION}"
    meta["A2"] = "Company"
    meta["B2"] = company
    meta["A3"] = "Generated at"
    meta["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta["A4"] = "Total profiles"
    meta["B4"] = len(profiles)
    meta["A5"] = "Payroll cross-referenced"
    meta["B5"] = "yes" if has_payroll else "no"
    meta.sheet_state = "hidden"   # don't distract HR from the main sheet

    wb.save(out)
    return len(profiles)

