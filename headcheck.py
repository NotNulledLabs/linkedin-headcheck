"""
LinkedIn HeadCheck
==================
Workforce verification tool — audits a LinkedIn company "People" page
snapshot and produces three output files:

  1. headcheck_<company>_<date>.html     — interactive report (all profiles)
  2. headcheck_<company>_<date>.pdf      — executive PDF report (all profiles)
  3. headcheck_<company>_suspects_<date>.csv — high/medium risk only (for HR meetings)

Usage
-----
  python headcheck.py --html people.html --company "Acme Corp"
  python headcheck.py --html people.html --company "Acme Corp" --payroll staff.xlsx
  python headcheck.py --html people.html --company "Acme Corp" --payroll staff.csv --lang es --out ./reports

How to export the LinkedIn People page
---------------------------------------
Option A — Browser Console (works in all browsers):
  1. Go to the company LinkedIn page → click "People" tab.
  2. Scroll all the way to the bottom so every profile card loads.
  3. Press F12 → click the Console tab.
  4. Paste:  copy(document.documentElement.outerHTML)  and press Enter.
  5. Open any text editor, paste (Ctrl+V), save as people.html.

Option B — Bookmarklet (more user-friendly, no console needed):
  See README.md for bookmarklet installation instructions.

Payroll file format
--------------------
  Accepts .csv, .xls, .xlsx — any layout.
  The script auto-detects the column containing employee names.
  No reformatting of your HR export is required.

Risk scoring (0–10)
--------------------
  +2  Has a loaded profile photo    (loaded = we actually see the CDN URL)
  +1  Photo exists but was not loaded at export time (lazy-load placeholder)
  +2  Has a custom URL slug         (not an auto-generated ACoAAA… ID)
  +1  Has a headline / job title
  +1  Headline references employer  ("at …" / "en …" / …)
  +2 / +3  Mutual connections       (1 mutual = +2, multiple = +3)
  −2  Suspicious name pattern       (initials only, numeric IDs, no spaces, etc.)

  Score >= 7  → Green  — complete, active profile
  Score  4-6  → Yellow — partial profile, review recommended
  Score <= 3  → Red    — ghost / minimal profile, HR priority

  Only a genuinely absent photo (no <img> or empty src) forces red. A photo
  that simply didn't load during export (ghost-person placeholder) is NOT
  penalised — it's a data-quality issue on the export side, not a property
  of the profile. Likewise, zero mutual connections no longer subtract from
  the score, because mutuals reflect the EXPORTER's network rather than the
  profile being analysed.
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from bs4 import BeautifulSoup

VERSION = "1.6.0"
TOOL_NAME = "LinkedIn HeadCheck"
BRAND_NAME = "Not Nulled Labs"
BRAND_URL  = "https://notnulled.com/"
REPO_URL   = "https://github.com/NotNulledLabs"

MAX_SCORE = 10


# ─────────────────────────────────────────────────────────────────────────────
# MUTUAL CONNECTION LANGUAGE PATTERNS
# ─────────────────────────────────────────────────────────────────────────────

MUTUAL_PATTERNS = {
    "en": re.compile(r"mutual connection", re.I),
    "es": re.compile(r"conexi[oó]n en com[uú]n|conexiones en com[uú]n", re.I),
    "fr": re.compile(r"relation commune|relations communes", re.I),
    "de": re.compile(r"gemeinsame.{0,6}kontakt", re.I),
    "pt": re.compile(r"conex[aã]o em comum|conex[oõ]es em comum", re.I),
    "it": re.compile(r"collegamento in comune|collegamenti in comune", re.I),
}

# Count any run of digits followed by a language-agnostic "mutual" keyword.
# We match the digit and let the per-language pattern confirm the context.
_DIGIT_PATTERN = re.compile(r"\b(\d+)\b")

# Conjunctions used before the "mutual" keyword to indicate 2+ shared contacts
# ("Bill and 2 others are mutual connections", "Ana y 3 más en común", …).
_MULTI_CONJUNCTIONS = re.compile(r"\b(and|y|et|und|e)\b", re.I)


def count_mutual(text: str, lang: str = "en") -> int:
    """
    Returns 0 (none), 1 (one mention), or 2 (multiple / explicit count >= 2).
    Checks the requested language pattern plus English as fallback.
    """
    if not text:
        return 0

    # Pick the language-specific pattern (+ English fallback) and locate the
    # position where the "mutual" keyword starts, if any.
    patterns = [p for p in (MUTUAL_PATTERNS.get(lang), MUTUAL_PATTERNS["en"]) if p]
    match = next((m for m in (p.search(text) for p in patterns) if m), None)
    if not match:
        return 0

    # Explicit numeric count: "3 mutual connections", "3 conexiones en común"…
    # We only look for digits that appear *before* the keyword to avoid
    # picking up unrelated numbers later in the string.
    prefix = text[: match.start()]
    digit_match = _DIGIT_PATTERN.search(prefix)
    if digit_match and int(digit_match.group(1)) >= 2:
        return 2

    # Conjunction fallback: "Bill and Sarah are mutual connections" →
    # two or more names joined before the keyword.
    if _MULTI_CONJUNCTIONS.search(prefix):
        return 2

    return 1


# ─────────────────────────────────────────────────────────────────────────────
# SUSPICIOUS NAME DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def is_suspicious_name(name: str) -> tuple[bool, str]:
    """
    Returns (is_suspicious, reason).
    Flags obvious anomalies without penalising non-English names.
    """
    if not name:
        return True, "empty name"

    # All uppercase with no spaces (e.g. "JOHNSMITHACCOUNT")
    if name == name.upper() and " " not in name and len(name) > 4:
        return True, "all-caps no spaces"

    # Contains digits (e.g. "John Smith123", "user_4892")
    if re.search(r"\d", name):
        return True, "contains digits"

    # Contains @ or common ID characters
    if re.search(r"[@_]", name):
        return True, "contains @ or _"

    # Only initials: "A. B." or "J.K." or single letter words only
    tokens = name.split()
    if tokens and all(re.match(r"^[A-Za-z]\.?$", t) for t in tokens):
        return True, "initials only"

    # Single token shorter than 3 chars (no surname).
    # Exception: CJK names are commonly 2 characters (e.g. 李明) and each
    # character carries meaning, so we don't flag them here.
    if len(tokens) == 1 and len(tokens[0]) < 3:
        if not re.search(r"[\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]", tokens[0]):
            return True, "single very short token"

    # No letters at all — covers Latin, Cyrillic, and common CJK ranges
    # (Chinese, hiragana/katakana, hangul) so non-Latin names aren't flagged.
    if not re.search(
        r"[A-Za-zÀ-ÿ\u0400-\u04FF\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]",
        name,
    ):
        return True, "no alphabetic characters"

    return False, ""


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def _clean_url(raw: str) -> str:
    url = raw.split("?")[0]
    # Strip any HTML entity artefacts that LinkedIn injects
    url = re.sub(r'["\'>]', "", url)
    url = url.replace("&amp;", "&").strip()
    return url


# LinkedIn uses a 1x1 transparent GIF as lazy-load placeholder when the image
# is still in the lazy-loading state (never in the viewport during export).
_LINKEDIN_GHOST_B64 = "data:image/gif;base64,R0lGODlhAQABAIA"

# Photo presence states:
#   "loaded"     — real photo URL captured; we have actual signal
#   "not_loaded" — lazy-load placeholder (ghost-person); UNKNOWN, not a red flag
#   "absent"     — no <img> tag or empty src; profile genuinely has no photo
PHOTO_LOADED = "loaded"
PHOTO_NOT_LOADED = "not_loaded"
PHOTO_ABSENT = "absent"


def _classify_photo(img_tag) -> tuple[str, str]:
    """
    Inspect an <img> tag and return (photo_state, avatar_url).

    `photo_state` is one of PHOTO_LOADED / PHOTO_NOT_LOADED / PHOTO_ABSENT.
    `avatar_url`  is the cleaned CDN URL when loaded, otherwise "".
    """
    if img_tag is None:
        return PHOTO_ABSENT, ""

    # The strongest signal LinkedIn gives us: the CSS class `ghost-person` is
    # applied to the <img> specifically when the picture hasn't hydrated yet.
    # This is a viewport/lazy-loading state, NOT "this person has no photo".
    classes = img_tag.get("class") or []
    if isinstance(classes, str):
        classes = classes.split()
    if any("ghost-person" in c for c in classes):
        return PHOTO_NOT_LOADED, ""

    raw = img_tag.get("src", "") or ""
    if not raw:
        return PHOTO_ABSENT, ""

    # Secondary heuristic: the 1x1 GIF data URL is the lazy placeholder source.
    # If we see it WITHOUT a ghost-person class, we still treat it as not-loaded
    # because that's what the byte-pattern means.
    if raw.startswith(_LINKEDIN_GHOST_B64):
        return PHOTO_NOT_LOADED, ""

    # Strip HTML entity artefacts LinkedIn injects in CDN URLs.
    url = re.sub(r'["\'>]+', "", raw)
    url = url.replace("&amp;", "&").strip()
    if not url.startswith(("http://", "https://")):
        return PHOTO_ABSENT, ""

    return PHOTO_LOADED, url


# Fallback-style field extraction: each field has a list of strategies
# ordered from most specific to most forgiving. If LinkedIn renames a
# CSS class in a UI refresh, a weaker strategy should still find the text
# so we don't silently return empty profiles.

def _find_name(card) -> str:
    # 1. The current (Jan 2026) class used by LinkedIn on People cards.
    for selector_fn in (
        lambda c: c.find("div", class_=re.compile(r"t-black")),
        # 2. Common ancestor class for profile cards — the title inside it.
        lambda c: c.find(class_=re.compile(r"artdeco-entity-lockup__title")),
        # 3. Any element that looks like a name anchor.
        lambda c: c.find("a", class_=re.compile(r"app-aware-link")),
    ):
        el = selector_fn(card)
        if el:
            txt = el.get_text(strip=True)
            if txt:
                return txt
    return ""


def _find_headline(card) -> str:
    for selector_fn in (
        lambda c: c.find("div", class_=re.compile(r"artdeco-entity-lockup__subtitle")),
        lambda c: c.find(class_=re.compile(r"entity-result__primary-subtitle")),
        lambda c: c.find(class_=re.compile(r"subtitle")),
    ):
        el = selector_fn(card)
        if el:
            txt = el.get_text(strip=True)
            if txt:
                return txt
    return ""


def _find_mutual_caption(card) -> str:
    for selector_fn in (
        lambda c: c.find("span", class_=re.compile(r"lt-line-clamp--multi-line t-12")),
        lambda c: c.find("span", class_=re.compile(r"entity-result__simple-insight-text")),
        lambda c: c.find(class_=re.compile(r"insight-text")),
    ):
        el = selector_fn(card)
        if el:
            txt = el.get_text(strip=True)
            if txt:
                return txt
    return ""


def extract_profiles(html_path: str, lang: str = "en",
                     debug: bool = False) -> tuple[list[dict], dict]:
    """
    Parse a LinkedIn People-page HTML snapshot.

    Returns (profiles, diagnostics) where diagnostics is a dict with counters
    the caller can use to warn the user about likely export quality issues
    (e.g. many un-loaded photos, suspiciously empty extractions).
    """
    with open(html_path, "r", encoding="utf-8", errors="replace") as f:
        soup = BeautifulSoup(f, "html.parser")

    anchors = soup.find_all(
        "a", id=re.compile(r"org-people-profile-card__profile-image-\d+")
    )

    people: list[dict] = []
    seen: set[str] = set()
    diag = {
        "anchors_found": len(anchors),
        "cards_skipped_no_parent": 0,
        "cards_skipped_no_name": 0,
        "cards_skipped_duplicate": 0,
        "photo_loaded": 0,
        "photo_not_loaded": 0,
        "photo_absent": 0,
    }

    for anchor in anchors:
        card = anchor.find_parent("div", class_="org-people-profile-card__profile-info")
        if not card:
            # Try a looser fallback for the card container before giving up.
            card = anchor.find_parent("div", class_=re.compile(r"profile-card|profile-info"))
            if not card:
                diag["cards_skipped_no_parent"] += 1
                continue

        name = _find_name(card)
        if not name:
            diag["cards_skipped_no_name"] += 1
            continue

        profile_url = _clean_url(anchor.get("href", ""))
        if not profile_url:
            continue
        if profile_url in seen:
            diag["cards_skipped_duplicate"] += 1
            continue
        seen.add(profile_url)

        photo_state, avatar_url = _classify_photo(anchor.find("img"))
        diag[f"photo_{photo_state}"] += 1

        headline = _find_headline(card)
        extra_note = _find_mutual_caption(card)

        slug = profile_url.replace("https://www.linkedin.com/in/", "").strip("/")
        slug_is_generic = bool(re.match(r"^ACo[A-Za-z0-9_\-]{10,}$", slug))
        mutual_level   = count_mutual(extra_note, lang)
        has_headline   = bool(headline)
        has_employer_ref = bool(
            headline and re.search(r"\bat\b|\ben\b|\bpri\b|\bbei\b|\bchez\b", headline, re.I)
        )
        susp_name, susp_reason = is_suspicious_name(name)

        profile = dict(
            name=name,
            profile_url=profile_url,
            avatar_url=avatar_url,
            photo_state=photo_state,              # new tri-state field
            has_photo=(photo_state == PHOTO_LOADED),  # kept for backward-compat
            headline=headline,
            extra_note=extra_note,
            slug=slug,
            slug_is_generic=slug_is_generic,
            mutual_level=mutual_level,
            has_headline=has_headline,
            has_employer_ref=has_employer_ref,
            suspicious_name=susp_name,
            suspicious_reason=susp_reason,
        )
        profile["score"] = _score(profile)
        profile["risk"]  = _risk(profile)
        people.append(profile)

    diag["profiles_extracted"] = len(people)
    # No printing from here — the diagnostics dict is returned so the caller
    # (CLI, TUI, tests…) decides whether and how to surface it. The `debug`
    # parameter is kept for backward compatibility.
    return people, diag


# ─────────────────────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────────────────────

def _score(p: dict) -> int:
    """
    Compute a 0–10 completeness score based on observable profile signals.

    Photo is the dominant signal: its absence (genuine, not "lazy-loading not
    finished") is still the single strongest risk indicator. But we no longer
    penalise profiles whose photo simply wasn't in the viewport at export
    time (see photo_state == PHOTO_NOT_LOADED), because that's a data-quality
    issue of the export, not a property of the profile itself.

    Similarly, zero mutual connections used to subtract a point. It no longer
    does: mutuals are a function of the exporter's network, not the profile's
    legitimacy. Missing evidence is not evidence against.
    """
    state = p.get("photo_state", PHOTO_LOADED if p.get("has_photo") else PHOTO_ABSENT)

    # Genuine absence: the profile really has no photo set. Still forces red.
    if state == PHOTO_ABSENT:
        base = 0
        if not p["slug_is_generic"]: base += 1
        if p["has_headline"]:        base += 1
        if p["mutual_level"] > 0:    base += 1
        return min(base, 3)

    # Photo loaded: full confidence in the signal.
    if state == PHOTO_LOADED:
        s = 2
    else:
        # Photo not loaded (ghost-person): unknown state. Give a single point
        # so the profile isn't punished for an export-side issue, but don't
        # grant the full photo bonus since we can't actually see it.
        s = 1

    if not p["slug_is_generic"]:  s += 2
    if p["has_headline"]:         s += 1
    if p["has_employer_ref"]:     s += 1

    # Mutual connections reinforce the signal when present. Their absence is
    # NOT a negative signal — it tells us about the exporter's network, not
    # about the profile being analysed.
    if p["mutual_level"] == 1:    s += 2
    elif p["mutual_level"] >= 2:  s += 3

    if p["suspicious_name"]:      s -= 2

    return max(0, min(s, MAX_SCORE))


def _risk(p: dict) -> str:
    # Only a genuinely absent photo forces red.
    state = p.get("photo_state", PHOTO_LOADED if p.get("has_photo") else PHOTO_ABSENT)
    if state == PHOTO_ABSENT:
        return "red"
    score = p["score"]
    if score >= 7: return "green"
    if score >= 4: return "yellow"
    return "red"


# ─────────────────────────────────────────────────────────────────────────────
# PAYROLL — flexible column detection
# ─────────────────────────────────────────────────────────────────────────────

# Preferred name-indicating keywords, ordered from strongest to weakest.
# "employee" / "staff" / "worker" match things like "Employee ID" which are
# NOT name columns, so we check the strong group first.
_NAME_HEADER_STRONG = re.compile(
    r"(full[\s_-]?name|\bname\b|\bnombre\b|\bapellido\b|\bnom\b|"
    r"mitarbeitername|nombre[\s_-]?completo)",
    re.I,
)
_NAME_HEADER_WEAK = re.compile(
    r"(empleado|employee|staff|worker|collaborator|collaborateur|"
    r"mitarbeiter|dipendente)",
    re.I,
)
# Kept for backward compatibility with older tests / callers.
_NAME_HEADER_RE = re.compile(
    _NAME_HEADER_STRONG.pattern + "|" + _NAME_HEADER_WEAK.pattern, re.I
)

def _col_name_score(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = sum(
        1 for v in values
        if v and re.match(r"^[A-Za-zÀ-ÿ'\-\. ]{3,60}$", v.strip())
        and len(v.strip().split()) >= 2
    )
    return hits / len(values)

def _best_name_col(rows: list[list[str]]) -> int:
    """
    Pick the column most likely to contain employee names.
    1. Prefer a column whose header strongly matches a name keyword
       ("Full Name", "Nombre"…).
    2. Accept a weak header match ("Employee", "Staff"…) only if its cells
       actually look like human names — otherwise it's probably an ID column.
    3. Fall back to the column whose cells best match a human-name pattern.
    """
    if not rows:
        return 0
    n_cols = max(len(r) for r in rows)
    header = rows[0]

    # Precompute per-column content scores once — reused by both strategies.
    content_scores = [
        _col_name_score([r[c] if c < len(r) else "" for r in rows[1:]])
        for c in range(n_cols)
    ]

    # 1. Strong header match wins immediately.
    for i in range(min(len(header), n_cols)):
        if _NAME_HEADER_STRONG.search(str(header[i])):
            return i

    # 2. Weak header match — only trust it if the column's content looks name-ish.
    for i in range(min(len(header), n_cols)):
        if _NAME_HEADER_WEAK.search(str(header[i])) and content_scores[i] >= 0.5:
            return i

    # 3. Pure content-based fallback.
    return content_scores.index(max(content_scores))

def load_payroll(path: str) -> list[str]:
    """
    Load names from a payroll file (.csv/.xlsx/.xls). Returns the detected
    name column as a list of strings. Silent by design — callers that want
    to log which column was auto-picked should use load_payroll_detailed().
    """
    names, _col_idx = load_payroll_detailed(path)
    return names


def load_payroll_detailed(path: str) -> tuple[list[str], int]:
    """
    Same as load_payroll but also returns the detected column index, so UIs
    can display that diagnostic to the user.
    """
    ext = Path(path).suffix.lower()
    rows: list[list[str]] = []

    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            for row in csv.reader(f):
                rows.append([str(c).strip() for c in row])
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    else:
        sys.exit(f"Unsupported payroll format: {ext}  (use .csv, .xlsx or .xls)")

    if not rows:
        return [], 0

    col   = _best_name_col(rows)
    start = 1 if (_NAME_HEADER_RE.search(rows[0][col] if col < len(rows[0]) else "")) else 0
    names = [r[col] for r in rows[start:] if col < len(r) and r[col]]
    return names, col


# ─────────────────────────────────────────────────────────────────────────────
# PAYROLL CROSS-REFERENCE
# ─────────────────────────────────────────────────────────────────────────────

def cross_reference(profiles: list[dict], payroll: list[str]) -> list[dict]:
    """
    Match each profile name against the payroll list.
    - Exact match is O(1) via a lowercased lookup dict.
    - Fuzzy match uses thefuzz.process.extractOne, which short-circuits on
      low-potential candidates and is noticeably faster than manually looping
      token_sort_ratio over every payroll row.
    """
    from thefuzz import fuzz, process

    # Build exact-match index once. Keeps the first occurrence if duplicates.
    exact_index: dict[str, str] = {}
    for pn in payroll:
        key = pn.lower().strip()
        if key and key not in exact_index:
            exact_index[key] = pn

    for p in profiles:
        nl = (p.get("name") or "").lower().strip()
        if not nl:
            p["payroll_status"] = "not_found"
            p["payroll_match"] = ""
            continue

        if nl in exact_index:
            p["payroll_status"] = "exact"
            p["payroll_match"] = exact_index[nl]
            continue

        result = process.extractOne(
            nl, payroll, scorer=fuzz.token_sort_ratio, score_cutoff=85
        )
        if result:
            p["payroll_status"] = "fuzzy"
            p["payroll_match"] = result[0]
        else:
            p["payroll_status"] = "not_found"
            p["payroll_match"] = ""

    return profiles


# ─────────────────────────────────────────────────────────────────────────────
# SUSPECTS CSV
# ─────────────────────────────────────────────────────────────────────────────

def export_suspects_csv(profiles: list[dict], has_payroll: bool, out: str) -> int:
    """
    Write the red+yellow subset to `out` as CSV. Returns the number of
    suspect profiles written.
    """
    suspects = [p for p in profiles if p["risk"] in ("red", "yellow")]
    suspects.sort(key=lambda p: p["score"])  # lowest score first

    fieldnames = ["risk", "score", "name", "headline", "profile_url",
                  "has_photo", "mutual_level", "slug_is_generic",
                  "suspicious_name", "suspicious_reason"]
    if has_payroll:
        fieldnames += ["payroll_status", "payroll_match"]

    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(suspects)

    return len(suspects)


def export_snapshot_json(profiles: list[dict], company: str,
                          has_payroll: bool, stats: dict, out: str) -> int:
    """
    Write a structured JSON snapshot of the audit for later diffing.

    Unlike the HTML / PDF / XLSX reports — which are meant for humans —
    this file is the durable machine-readable record of what was observed
    on this date. Two snapshots can be compared with `diff_snapshots()` to
    answer questions like "who is new since last audit?".

    Returns the number of profile entries serialised.
    """
    payload = {
        "headcheck_version": VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "company": company,
        "has_payroll": has_payroll,
        "stats": stats,
        "profiles": profiles,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return len(profiles)


# ─────────────────────────────────────────────────────────────────────────────
# SNAPSHOT DIFF
# ─────────────────────────────────────────────────────────────────────────────

def _load_snapshot(path: str) -> dict:
    """Load and lightly validate a HeadCheck JSON snapshot."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "profiles" not in data:
        sys.exit(f"{path!r} does not look like a HeadCheck snapshot "
                 "(missing 'profiles' key).")
    return data


def diff_snapshots(old_path: str, new_path: str) -> dict:
    """
    Compare two HeadCheck snapshots and classify profiles by what changed.

    Identity is established by `profile_url`: if LinkedIn re-uses the same
    URL slug, it's the same person. A profile whose slug changed will show
    up as both "appeared" and "disappeared" — an inherent limitation of
    using public data without LinkedIn's internal IDs.

    Returns a dict with these keys, each mapping to a list of profiles:
        appeared:       profiles in new but not in old
        disappeared:    profiles in old but not in new
        risk_up:        risk got worse (green→yellow, yellow→red, green→red)
        risk_down:      risk got better
        score_changed:  score changed but risk stayed the same (minor drift)
        unchanged:      same in every relevant way
    """
    old = _load_snapshot(old_path)
    new = _load_snapshot(new_path)

    old_by_url = {p["profile_url"]: p for p in old["profiles"] if p.get("profile_url")}
    new_by_url = {p["profile_url"]: p for p in new["profiles"] if p.get("profile_url")}

    old_urls = set(old_by_url)
    new_urls = set(new_by_url)

    result = {
        "appeared":      [new_by_url[u] for u in sorted(new_urls - old_urls)],
        "disappeared":   [old_by_url[u] for u in sorted(old_urls - new_urls)],
        "risk_up":       [],
        "risk_down":     [],
        "score_changed": [],
        "unchanged":     [],
        "meta": {
            "old_path": old_path,
            "new_path": new_path,
            "old_generated_at": old.get("generated_at"),
            "new_generated_at": new.get("generated_at"),
            "old_company": old.get("company"),
            "new_company": new.get("company"),
        },
    }

    # Order risk states from safest to most concerning so we can compare.
    risk_rank = {"green": 0, "yellow": 1, "red": 2}

    for url in sorted(old_urls & new_urls):
        op, np = old_by_url[url], new_by_url[url]
        old_rank = risk_rank.get(op.get("risk"), 1)
        new_rank = risk_rank.get(np.get("risk"), 1)
        entry = {
            "name":        np.get("name") or op.get("name"),
            "profile_url": url,
            "old_risk":    op.get("risk"),
            "new_risk":    np.get("risk"),
            "old_score":   op.get("score"),
            "new_score":   np.get("score"),
        }
        if new_rank > old_rank:
            result["risk_up"].append(entry)
        elif new_rank < old_rank:
            result["risk_down"].append(entry)
        elif op.get("score") != np.get("score"):
            result["score_changed"].append(entry)
        else:
            result["unchanged"].append(entry)

    return result


def export_diff_csv(diff: dict, out: str) -> int:
    """
    Flatten the diff result into a single CSV for HR spreadsheets.
    Returns the number of rows written (excluding 'unchanged' profiles,
    which would be noise in a diff report).
    """
    rows = []
    for p in diff["appeared"]:
        rows.append({
            "change": "appeared",
            "name": p.get("name", ""),
            "profile_url": p.get("profile_url", ""),
            "old_risk": "",
            "new_risk": p.get("risk", ""),
            "old_score": "",
            "new_score": p.get("score", ""),
        })
    for p in diff["disappeared"]:
        rows.append({
            "change": "disappeared",
            "name": p.get("name", ""),
            "profile_url": p.get("profile_url", ""),
            "old_risk": p.get("risk", ""),
            "new_risk": "",
            "old_score": p.get("score", ""),
            "new_score": "",
        })
    for category in ("risk_up", "risk_down", "score_changed"):
        for p in diff[category]:
            rows.append({
                "change": category,
                "name": p["name"],
                "profile_url": p["profile_url"],
                "old_risk": p["old_risk"],
                "new_risk": p["new_risk"],
                "old_score": p["old_score"],
                "new_score": p["new_score"],
            })

    fieldnames = ["change", "name", "profile_url",
                  "old_risk", "new_risk", "old_score", "new_score"]
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# XLSX REPORT
# ─────────────────────────────────────────────────────────────────────────────

# Row-level background colours for risk cells. Chosen light enough that the
# default black text remains readable, and distinct enough that someone
# skimming the sheet can scan by colour alone.
_XLSX_RISK_FILL = {
    "red":    "FCE4E4",   # light red
    "yellow": "FFF4C2",   # light amber
    "green":  "DCEEDC",   # light green
}

# Risk labels are kept human-readable in the sheet. The raw code is available
# elsewhere via the 'risk' column, but labels are what HR will scan.
_XLSX_RISK_LABEL = {
    "red":    "🔴 High risk",
    "yellow": "🟡 Needs review",
    "green":  "🟢 Low risk",
}


def generate_xlsx(profiles: list[dict], company: str, has_payroll: bool,
                  out: str) -> int:
    """
    Write the full profile list to an Excel workbook tuned for HR review:
    frozen header row, autofilter, colour-coded risk column, score data
    bars, hyperlinked profile URLs, and an empty Notes column for inline
    annotation. Returns the number of profile rows written.

    This report is complementary to the suspects CSV (red+yellow only) and
    to the interactive HTML (all profiles, with search/filter in-browser).
    The XLSX exists for HR teams who work in Excel end-to-end: they can
    sort, filter, and annotate without leaving their usual tool.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HeadCheck"

    # ── Header row ──
    headers = [
        "Risk", "Score", "Name", "Headline", "Profile URL",
        "Photo state", "Mutual level", "Generic slug?",
        "Suspicious name?", "Suspicious reason",
    ]
    if has_payroll:
        headers += ["Payroll status", "Payroll match"]
    headers += ["HR notes"]

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A2F3A")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DE"),
        right=Side(style="thin", color="D5D8DE"),
        top=Side(style="thin", color="D5D8DE"),
        bottom=Side(style="thin", color="D5D8DE"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = thin_border

    # ── Data rows, ordered worst-first so the interesting ones appear on top ──
    ordered = sorted(profiles, key=lambda p: (p["score"], p["name"].lower()))

    for p in ordered:
        row = [
            _XLSX_RISK_LABEL.get(p["risk"], p["risk"]),
            p["score"],
            p["name"],
            p.get("headline", ""),
            p.get("profile_url", ""),
            p.get("photo_state", "loaded" if p.get("has_photo") else "absent"),
            p.get("mutual_level", 0),
            "yes" if p.get("slug_is_generic") else "no",
            "yes" if p.get("suspicious_name") else "no",
            p.get("suspicious_reason", ""),
        ]
        if has_payroll:
            row += [p.get("payroll_status", ""), p.get("payroll_match", "")]
        row += [""]   # empty Notes column for HR to fill in

        ws.append(row)
        r = ws.max_row

        # Risk cell background — makes the sheet scannable at a glance.
        ws.cell(row=r, column=1).fill = PatternFill(
            "solid", fgColor=_XLSX_RISK_FILL.get(p["risk"], "FFFFFF")
        )

        # Profile URL as hyperlink — HR can click through from the sheet.
        url_cell = ws.cell(row=r, column=5)
        if p.get("profile_url"):
            url_cell.hyperlink = p["profile_url"]
            url_cell.font = Font(color="2563EB", underline="single")

    # ── Visuals: data bars on the score column ──
    # Score column is #2; data rows start at row 2. Use row count for range.
    last_row = ws.max_row
    if last_row >= 2:
        score_range = f"B2:B{last_row}"
        ws.conditional_formatting.add(
            score_range,
            DataBarRule(start_type="num", start_value=0,
                        end_type="num", end_value=MAX_SCORE,
                        color="4CAF50", showValue=True),
        )

    # ── Freeze header + enable autofilter so HR can sort/filter ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Column widths: set once based on header length + generous padding.
    # Fine-tuning is manual work HR can do from Excel itself if needed.
    widths = {
        "A": 18,   # Risk
        "B": 8,    # Score
        "C": 28,   # Name
        "D": 50,   # Headline
        "E": 45,   # URL
        "F": 13,   # Photo state
        "G": 14,   # Mutual level
        "H": 14,   # Generic slug?
        "I": 17,   # Suspicious name?
        "J": 28,   # Suspicious reason
    }
    if has_payroll:
        widths.update({"K": 16, "L": 28})                                 # payroll cols
        widths[get_column_letter(13)] = 40                                # notes
    else:
        widths[get_column_letter(11)] = 40                                # notes
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # A tiny metadata sheet — for future diffs and provenance.
    meta = wb.create_sheet("_meta")
    meta["A1"] = "Generated by"
    meta["B1"] = f"LinkedIn HeadCheck v{VERSION}"
    meta["A2"] = "Company"
    meta["B2"] = company
    meta["A3"] = "Generated at"
    meta["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta["A4"] = "Total profiles"
    meta["B4"] = len(profiles)
    meta["A5"] = "Payroll cross-referenced"
    meta["B5"] = "yes" if has_payroll else "no"
    meta.sheet_state = "hidden"   # don't distract HR from the main sheet

    wb.save(out)
    return len(profiles)


# ─────────────────────────────────────────────────────────────────────────────
# HTML REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_html(profiles: list[dict], company: str, has_payroll: bool, out: str):
    total  = len(profiles)
    green  = sum(1 for p in profiles if p["risk"] == "green")
    yellow = sum(1 for p in profiles if p["risk"] == "yellow")
    red    = sum(1 for p in profiles if p["risk"] == "red")
    ts     = datetime.now().strftime("%B %d, %Y — %H:%M")
    # Stable key for scoping localStorage notes per company — must match main()'s slug.
    co_slug = re.sub(r"[^a-z0-9]+", "_", company.lower()).strip("_") or "company"

    if has_payroll:
        exact     = sum(1 for p in profiles if p.get("payroll_status") == "exact")
        fuzzy_n   = sum(1 for p in profiles if p.get("payroll_status") == "fuzzy")
        not_found = sum(1 for p in profiles if p.get("payroll_status") == "not_found")
    else:
        exact = fuzzy_n = not_found = 0

    # Escape sequences that could break out of the surrounding <script> tag.
    # json.dumps doesn't escape < / > / & by default, so a profile name
    # containing "</script>" would terminate the script early. We replace
    # those characters with their \u-escaped forms, which remain valid JSON
    # but are inert inside an HTML document.
    data_json = (
        json.dumps(profiles, ensure_ascii=False)
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("&", "\\u0026")
        .replace("\u2028", "\\u2028")  # JS line terminators break JSON in <script>
        .replace("\u2029", "\\u2029")
    )

    payroll_th  = "<th>Payroll <span class='tip' data-tip='Whether this name was found in the uploaded payroll file.'>?</span></th><th>Matched name</th>" if has_payroll else ""
    payroll_sel = """<select id="fPayroll">
          <option value="">All payroll statuses</option>
          <option value="exact">In payroll</option>
          <option value="fuzzy">Similar name</option>
          <option value="not_found">Not found</option>
        </select>""" if has_payroll else ""

    payroll_kpis = f"""
      <div class="kpi"><div class="kpi-val c-green">{exact}</div><div class="kpi-lbl">In payroll</div></div>
      <div class="kpi"><div class="kpi-val c-amber">{fuzzy_n}</div><div class="kpi-lbl">Similar name</div></div>
      <div class="kpi"><div class="kpi-val c-red">{not_found}</div><div class="kpi-lbl">Not in payroll</div></div>
    """ if has_payroll else ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>LinkedIn HeadCheck — {company}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

:root {{
  --bg:        #F2F4F8;
  --surface:   #FFFFFF;
  --border:    #D9DEE9;
  --text:      #111827;
  --muted:     #4B5563;
  --blue:      #1D4ED8;
  --blue-lt:   #EFF6FF;
  --blue-dk:   #1E3A8A;
  --green:     #166534;
  --green-lt:  #DCFCE7;
  --amber:     #92400E;
  --amber-lt:  #FEF3C7;
  --red:       #991B1B;
  --red-lt:    #FEE2E2;
  --radius:    8px;
  --shadow:    0 1px 3px rgba(0,0,0,.10), 0 1px 2px rgba(0,0,0,.06);
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Inter', sans-serif; background: var(--bg); color: var(--text); font-size: 14px; line-height: 1.5; }}

/* HEADER */
.header {{ background: var(--blue-dk); color: #fff; padding: 0; }}
.header-inner {{ padding: 24px 40px 20px; }}
.brand-row {{ display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px; margin-bottom: 14px; }}
.brand {{ display: flex; align-items: center; gap: 10px; }}
.brand-icon {{ width: 34px; height: 34px; background: #fff; border-radius: 6px; display: flex; align-items: center; justify-content: center; color: var(--blue-dk); font-weight: 800; font-size: 13px; font-family: 'IBM Plex Mono', monospace; flex-shrink: 0; }}
.brand-name {{ font-size: 13px; font-weight: 600; color: rgba(255,255,255,.9); }}
.brand-sub  {{ font-size: 11px; color: rgba(255,255,255,.55); }}
.header-meta {{ font-size: 11px; color: rgba(255,255,255,.55); text-align: right; line-height: 1.7; }}
.header h1 {{ font-size: 20px; font-weight: 700; color: #fff; }}
.header h1 span {{ color: #93C5FD; }}

/* HELP PANEL */
.help-bar {{
  background: #1E40AF;
  border-bottom: 1px solid #1D4ED8;
  padding: 0 40px;
  display: flex; gap: 0; overflow: hidden;
}}
.help-tab {{
  padding: 10px 18px;
  font-size: 13px; font-weight: 500;
  color: rgba(255,255,255,.65);
  cursor: pointer; border: none; background: none;
  border-bottom: 2px solid transparent;
  transition: color .15s, border-color .15s;
}}
.help-tab.active {{ color: #fff; border-bottom-color: #93C5FD; }}
.help-panel {{
  background: #1E3A8A;
  border-bottom: 1px solid #1D4ED8;
  padding: 20px 40px;
  font-size: 14px; color: rgba(255,255,255,.85);
  line-height: 1.8;
  display: none;
}}
.help-panel.active {{ display: block; }}
.help-panel h3 {{ font-size: 13px; font-weight: 700; color: #fff; margin-bottom: 10px; text-transform: uppercase; letter-spacing: .06em; }}
.help-panel p {{ margin-bottom: 10px; }}
.help-panel code {{ background: rgba(255,255,255,.12); padding: 2px 7px; border-radius: 4px; font-family: 'IBM Plex Mono', monospace; font-size: 12px; }}
.help-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }}
.help-grid dt {{ font-weight: 600; color: #93C5FD; margin-bottom: 4px; font-size: 13px; }}
.help-grid dd {{ color: rgba(255,255,255,.8); font-size: 13px; line-height: 1.6; }}

/* KPI BAR */
.kpi-bar {{ background: var(--surface); border-bottom: 1px solid var(--border); padding: 14px 40px; display: flex; gap: 10px; flex-wrap: wrap; box-shadow: var(--shadow); }}
.kpi {{ background: var(--bg); border: 1px solid var(--border); border-radius: var(--radius); padding: 12px 18px; min-width: 96px; text-align: center; }}
.kpi-val {{ font-size: 24px; font-weight: 700; line-height: 1; font-family: 'IBM Plex Mono', monospace; }}
.kpi-lbl {{ font-size: 10px; color: var(--muted); margin-top: 3px; text-transform: uppercase; letter-spacing: .05em; }}
.c-blue  {{ color: var(--blue);  }}
.c-green {{ color: var(--green); }}
.c-amber {{ color: var(--amber); }}
.c-red   {{ color: var(--red);   }}

/* FILTERS */
.filters {{
  position: sticky; top: 0; z-index: 60;
  background: var(--surface);
  border-bottom: 1px solid var(--border);
  padding: 10px 40px;
  display: flex; gap: 8px; flex-wrap: wrap; align-items: center;
  box-shadow: var(--shadow);
}}
.filters input, .filters select {{
  border: 1px solid var(--border); border-radius: var(--radius);
  padding: 7px 11px; font-family: 'Inter', sans-serif; font-size: 13px;
  color: var(--text); background: var(--bg); outline: none;
  transition: border-color .15s;
}}
.filters input {{ min-width: 210px; }}
.filters input:focus, .filters select:focus {{ border-color: var(--blue); }}
.filter-info {{ margin-left: auto; font-size: 11px; color: var(--muted); font-family: 'IBM Plex Mono', monospace; }}
.btn-export {{
  margin-left: 8px;
  padding: 7px 14px;
  background: var(--blue); color: #fff;
  border: none; border-radius: var(--radius);
  font-size: 12px; font-weight: 600; font-family: 'Inter', sans-serif;
  cursor: pointer; transition: background .15s;
}}
.btn-export:hover {{ background: var(--blue-dk); }}

/* TABLE */
.table-wrap {{ padding: 20px 40px 48px; overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; background: var(--surface); border-radius: var(--radius); overflow: hidden; box-shadow: var(--shadow); }}
thead th {{
  background: #F8FAFD; border-bottom: 2px solid var(--border);
  padding: 10px 13px; text-align: left;
  font-size: 11px; font-weight: 600; color: var(--muted);
  text-transform: uppercase; letter-spacing: .06em;
  white-space: nowrap;
}}
tbody tr {{ border-bottom: 1px solid var(--border); transition: background .1s; }}
tbody tr:last-child {{ border-bottom: none; }}
tbody tr:hover {{ background: #F8FAFF; }}
tbody td {{ padding: 10px 13px; vertical-align: middle; }}

/* AVATAR */
.av-cell {{ width: 44px; }}
.av {{ width: 38px; height: 38px; border-radius: 50%; object-fit: cover; border: 2px solid var(--border); display: block; }}
.av-wrap {{ display: inline-block; }}
.av-ph {{ width: 38px; height: 38px; border-radius: 50%; background: #E5E7EB; display: flex; align-items: center; justify-content: center; font-size: 16px; border: 2px dashed #D1D5DB; color: #9CA3AF; }}

/* BADGES */
.badge {{ display: inline-block; border-radius: 4px; padding: 3px 9px; font-size: 11px; font-weight: 600; white-space: nowrap; }}
.badge-green  {{ background: var(--green-lt);  color: var(--green);  }}
.badge-yellow {{ background: var(--amber-lt);  color: var(--amber);  }}
.badge-red    {{ background: var(--red-lt);    color: var(--red);    }}
.badge-susp   {{ background: #FEF9C3; color: #713F12; font-size: 10px; margin-left: 4px; }}

/* SCORE */
.sc-wrap {{ display: flex; align-items: center; gap: 6px; }}
.sc-bar  {{ width: 52px; height: 5px; background: #E5E7EB; border-radius: 3px; overflow: hidden; }}
.sc-fill {{ height: 100%; border-radius: 3px; }}
.sc-txt  {{ font-size: 11px; color: var(--muted); font-family: 'IBM Plex Mono', monospace; min-width: 28px; }}

/* INDICATORS */
.inds {{ display: flex; gap: 2px; flex-wrap: wrap; }}
.ind     {{ font-size: 14px; cursor: default; }}
.ind-off {{ opacity: .2; filter: grayscale(1); }}

/* HEADLINE */
.hl-cell {{ max-width: 220px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; color: var(--muted); font-size: 12px; }}

/* LINK */
a.pl {{ color: var(--blue); text-decoration: none; font-size: 12px; font-weight: 500; }}
a.pl:hover {{ text-decoration: underline; }}

/* PAYROLL */
.ps {{ font-size: 11px; font-weight: 600; }}
.ps-exact     {{ color: var(--green); }}
.ps-fuzzy     {{ color: var(--amber); }}
.ps-not_found {{ color: var(--red);   }}
.pm {{ font-size: 11px; color: var(--muted); max-width: 140px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; display: block; }}

/* TOOLTIP */
.tip {{
  display: inline-flex; align-items: center; justify-content: center;
  width: 15px; height: 15px; border-radius: 50%;
  background: var(--border); color: var(--muted);
  font-size: 10px; font-weight: 700; cursor: help;
  position: relative; vertical-align: middle; margin-left: 3px;
}}
.tip:hover::after {{
  content: attr(data-tip);
  position: absolute; bottom: 120%; left: 50%; transform: translateX(-50%);
  background: #1F2937; color: #fff;
  padding: 6px 10px; border-radius: 6px;
  font-size: 11px; font-weight: 400; white-space: nowrap;
  z-index: 999; pointer-events: none;
  box-shadow: 0 2px 8px rgba(0,0,0,.2);
}}

/* NO RESULTS */
.no-res {{ display: none; text-align: center; padding: 48px; color: var(--muted); font-size: 13px; }}

/* NOTES */
.note-input {{
  border: 1px solid var(--border); border-radius: 4px;
  padding: 4px 7px; font-size: 11px; font-family: 'Inter', sans-serif;
  color: var(--text); background: var(--bg); width: 120px;
  transition: border-color .15s;
}}
.note-input:focus {{ outline: none; border-color: var(--blue); }}

/* FOOTER */
footer {{
  border-top: 1px solid var(--border);
  padding: 14px 40px;
  font-size: 11px; color: var(--muted);
  display: flex; justify-content: space-between; flex-wrap: wrap; gap: 8px;
  background: var(--surface);
}}
footer a {{ color: var(--blue); text-decoration: none; }}
footer a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>

<!-- HEADER -->
<div class="header">
  <div class="header-inner">
    <div class="brand-row">
      <div class="brand">
        <div class="brand-icon">HC</div>
        <div>
          <div class="brand-name">LinkedIn HeadCheck</div>
          <div class="brand-sub">Security &amp; Workforce Verification Tool by Not Nulled Labs</div>
        </div>
      </div>
      <div class="header-meta">
        Generated: {ts}<br>
        Profiles analysed: <strong style="color:#fff">{total}</strong>
      </div>
    </div>
    <h1>People Audit — <span>{company}</span></h1>
  </div>

  <!-- HELP TABS -->
  <div class="help-bar">
    <button class="help-tab" onclick="toggleHelp('how', this)">How it works</button>
    <button class="help-tab" onclick="toggleHelp('scoring', this)">Scoring explained</button>
    <button class="help-tab" onclick="toggleHelp('glossary', this)">Glossary</button>
    <button class="help-tab" onclick="toggleHelp('nextsteps', this)" style="color:#86EFAC">Next steps</button>
  </div>

  <div class="help-panel" id="help-how">
    <h3>How This Report Was Generated</h3>
    <p>A LinkedIn company "People" page was exported as an HTML snapshot using the browser console command <code>copy(document.documentElement.outerHTML)</code>. LinkedIn HeadCheck then parsed that snapshot and scored each profile based on publicly visible signals.</p>
    <p>No scraping was performed. No LinkedIn API was used. The data reflects what was visible on screen at the time of export.</p>
    <p><strong>Important:</strong> Risk scores indicate profile completeness — not confirmed employment. A complete profile does not guarantee the person currently works here. All findings should be verified by HR against internal records.</p>
    <p style="margin-top:10px;padding:10px 14px;background:rgba(245,158,11,.12);border-left:3px solid #f59e0b;border-radius:4px;font-size:13px">
      <strong>⚠️ If you are a LinkedIn page admin:</strong> LinkedIn does not show your personal network in the admin view, even with <code>?viewAsMember=true</code>. This means mutual connections will be absent for all profiles, causing legitimate employees to appear as yellow instead of green. For accurate results, the export should be done by a non-admin employee navigating to the People page from their normal LinkedIn session.
    </p>
  </div>

  <div class="help-panel" id="help-scoring">
    <h3>How Risk Scores Work (0 – {MAX_SCORE})</h3>
    <div class="help-grid">
      <dl><dt>+2 — Profile photo</dt><dd>Has a visible profile picture. Absence forces the profile to Red regardless of other signals.</dd></dl>
      <dl><dt>+2 — Custom URL</dt><dd>Profile has a personalised slug (e.g. /in/john-smith). Auto-generated IDs like ACoAAA… suggest an unconfigured account.</dd></dl>
      <dl><dt>+1 — Headline present</dt><dd>Profile has a job title or headline text.</dd></dl>
      <dl><dt>+1 — Employer reference</dt><dd>Headline contains "at", "en", or similar linking the person to an employer.</dd></dl>
      <dl><dt>+2 or +3 — Mutual connections</dt><dd>LinkedIn shows connections shared with the company network. One mutual = +2, multiple = +3. Zero mutuals = −1 penalty.</dd></dl>
      <dl><dt>−2 — Suspicious name</dt><dd>Name contains digits, initials only, special characters, or other unusual patterns.</dd></dl>
    </div>
    <p style="margin-top:12px">Score ≥ 7 → Green &nbsp;·&nbsp; Score 4–6 → Yellow &nbsp;·&nbsp; Score ≤ 3 → Red</p>
    <p style="margin-top:10px;font-size:13px;opacity:.75">Profiles exported from a page admin account will show no mutual connections and will score lower as a result. This is a LinkedIn limitation.</p>
  </div>

  <div class="help-panel" id="help-glossary">
    <h3>Glossary</h3>
    <div class="help-grid">
      <dl><dt>Mutual connections</dt><dd>People in the LinkedIn network of the account used to export the page who are also connected to this profile. Not visible when exported from a page admin account.</dd></dl>
      <dl><dt>Custom URL</dt><dd>A profile address chosen by the user (e.g. /in/john-smith) as opposed to an auto-generated ID.</dd></dl>
      <dl><dt>Slug</dt><dd>The identifier part of a LinkedIn profile URL after /in/.</dd></dl>
      <dl><dt>Payroll match — Exact</dt><dd>The name appears in the payroll file with identical spelling.</dd></dl>
      <dl><dt>Payroll match — Similar</dt><dd>A close but not identical match was found (fuzzy matching, ≥85% similarity).</dd></dl>
      <dl><dt>Payroll match — Not found</dt><dd>No matching name found. May indicate a former employee, contractor, or error in either list.</dd></dl>
      <dl><dt>Suspicious name</dt><dd>The name contains digits, only initials, special characters, or other patterns uncommon in human names.</dd></dl>
      <dl><dt>Page admin limitation</dt><dd>LinkedIn page admins cannot view the People section with their personal network. Use a non-admin account for best results.</dd></dl>
    </div>
  </div>

  <div class="help-panel" id="help-nextsteps">
    <h3>What to Do After the Audit</h3>
    <p>Once you have identified suspicious profiles, these are the actions available to you depending on the case.</p>

    <div class="help-grid" style="margin-top:14px">
      <div>
        <dl><dt style="color:#86EFAC;margin-bottom:6px">👋 Former employee — forgot to update</dt>
        <dd>The most common case. Contact the person directly and ask them to edit the Experience section on their LinkedIn profile and remove or update the position. LinkedIn will automatically disassociate their profile. Changes can take up to 30 days to reflect.</dd></dl>
      </div>
      <div>
        <dl><dt style="color:#FCA5A5;margin-bottom:6px">🚨 Fake or malicious profile</dt>
        <dd>
          <strong>Step 1 — Report to LinkedIn</strong><br>
          Admins cannot remove members directly. Contact LinkedIn via their support form providing: the person's full name, a screenshot of the People page showing them, a link to their profile, and an explanation.<br>
          Form: <a href="https://www.linkedin.com/help/linkedin/ask/cp-master" target="_blank" style="color:#93C5FD">linkedin.com/help/linkedin/ask/cp-master</a><br><br>
          <strong>Step 2 — Report the profile</strong><br>
          On the person's LinkedIn profile, click ··· → Report / Block → select the appropriate reason (fake profile, incorrect information).<br><br>
          <strong>Step 3 — Escalate if needed</strong><br>
          If there is no response, contacting <a href="https://twitter.com/LinkedInHelp" target="_blank" style="color:#93C5FD">@LinkedInHelp</a> on X/Twitter can speed up the process.
        </dd></dl>
      </div>
    </div>

    <div style="margin-top:16px;padding:10px 14px;background:rgba(239,68,68,.1);border-left:3px solid #ef4444;border-radius:4px;font-size:11px">
      <strong>What admins cannot do:</strong> Admins cannot directly remove a member from the company page, cannot see a full list of associated members in the admin dashboard, and cannot prevent someone from self-associating with the company. LinkedIn's system relies entirely on self-reported data — which is exactly why this tool exists.
    </div>
  </div>
</div>

<!-- KPI BAR -->
<div class="kpi-bar">
  <div class="kpi"><div class="kpi-val c-blue">{total}</div><div class="kpi-lbl">Total</div></div>
  <div class="kpi"><div class="kpi-val c-green">{green}</div><div class="kpi-lbl">Low risk</div></div>
  <div class="kpi"><div class="kpi-val c-amber">{yellow}</div><div class="kpi-lbl">Review</div></div>
  <div class="kpi"><div class="kpi-val c-red">{red}</div><div class="kpi-lbl">High risk</div></div>
  {payroll_kpis}
</div>

<!-- FILTERS -->
<div class="filters">
  <input type="text" id="fSearch" placeholder="Search by name or headline…">
  <select id="fRisk">
    <option value="">All risk levels</option>
    <option value="green">Low risk</option>
    <option value="yellow">Needs review</option>
    <option value="red">High risk</option>
  </select>
  <select id="fMutual">
    <option value="">All connections</option>
    <option value="0">No mutual connections</option>
    <option value="1">1 mutual connection</option>
    <option value="2">Multiple mutual connections</option>
  </select>
  {payroll_sel}
  <button class="btn-export" onclick="exportNotes()">Export notes CSV</button>
  <span class="filter-info" id="fCount">{total} profiles shown</span>
</div>

<!-- TABLE -->
<div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th class="av-cell"></th>
        <th>Name <span class="tip" data-tip="Full name as shown on LinkedIn">?</span></th>
        <th>Headline <span class="tip" data-tip="Job title or headline from the profile">?</span></th>
        <th>Risk <span class="tip" data-tip="Green = low risk · Yellow = review needed · Red = high risk or no photo">?</span></th>
        <th>Score <span class="tip" data-tip="Completeness score from 0 to {MAX_SCORE}. Higher = more complete profile.">?</span></th>
        <th>Signals <span class="tip" data-tip="📷 photo · ✏️ headline · 🔗 custom URL · 🤝 mutual connections · ⚠️ suspicious name">?</span></th>
        {payroll_th}
        <th>Notes <span class="tip" data-tip="Add HR notes per profile. Export all notes with the button above.">?</span></th>
        <th>Profile</th>
      </tr>
    </thead>
    <tbody id="tbody"></tbody>
  </table>
  <div class="no-res" id="noRes">No profiles match the current filters.</div>
</div>

<!-- FOOTER -->
<footer>
  <span>
    <a href="{REPO_URL}" target="_blank" rel="noopener">LinkedIn HeadCheck</a>
    &nbsp;·&nbsp;
    <a href="{BRAND_URL}" target="_blank" rel="noopener">{BRAND_NAME}</a>
    &nbsp;·&nbsp; v{VERSION}
  </span>
  <span>{company} &nbsp;·&nbsp; {ts}</span>
</footer>

<script>
const profiles   = {data_json};
const hasPayroll = {'true' if has_payroll else 'false'};
const MAX        = {MAX_SCORE};

// Escape HTML special characters before interpolating any user-controlled
// string into innerHTML. LinkedIn profile text is ultimately attacker-
// controlled, so we never trust it.
const escapeHtml = (s) => String(s ?? '').replace(/[&<>"']/g, c => ({{
  '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'
}}[c]));

// Persist notes across page reloads so HR can pause and resume.
const NOTES_KEY = 'headcheck_notes_{co_slug}';
let notes = {{}};
try {{ notes = JSON.parse(localStorage.getItem(NOTES_KEY) || '{{}}'); }} catch (e) {{ notes = {{}}; }}
function saveNotes() {{
  try {{ localStorage.setItem(NOTES_KEY, JSON.stringify(notes)); }} catch (e) {{}}
}}

function scoreColor(s) {{
  if (s / MAX >= 0.7) return 'var(--green)';
  if (s / MAX >= 0.4) return 'var(--amber)';
  return 'var(--red)';
}}

const RISK_LABELS = {{ green:'Low risk', yellow:'Needs review', red:'High risk' }};
const PS_ICONS    = {{ exact:'&#10003; In payroll', fuzzy:'&#126; Similar', not_found:'&#10007; Not found' }};
const PS_CLS      = {{ exact:'ps-exact', fuzzy:'ps-fuzzy', not_found:'ps-not_found' }};
const MUTUAL_LABELS = ['No mutual', '1 mutual', 'Multiple mutuals'];

function ind(icon, active, tip) {{
  return `<span class="ind ${{active?'':'ind-off'}}" title="${{escapeHtml(tip)}}">${{icon}}</span>`;
}}

function payrollCells(p) {{
  if (!hasPayroll) return '';
  const st = p.payroll_status || 'not_found';
  const match = escapeHtml(p.payroll_match || '');
  return `<td><span class="ps ${{PS_CLS[st]||''}}">${{PS_ICONS[st]||'—'}}</span></td>`
       + `<td><span class="pm" title="${{match}}">${{match || '—'}}</span></td>`;
}}

function renderRow(p, idx) {{
  const safeName     = escapeHtml(p.name);
  const safeHeadline = escapeHtml(p.headline || '');
  const safeAvatar   = escapeHtml(p.avatar_url || '');
  const safeUrl      = escapeHtml(p.profile_url);
  const safeReason   = escapeHtml(p.suspicious_reason || '');

  // Render both img and placeholder — toggle visibility via onerror/onload
  // to avoid unescaped HTML chars inside onerror attribute breaking the parser
  const av = p.avatar_url
    ? `<span class="av-wrap">
        <img class="av" src="${{safeAvatar}}" alt=""
             loading="lazy"
             onerror="this.style.display='none';this.parentNode.querySelector('.av-ph').style.display='flex'">
        <span class="av-ph" style="display:none">&#128100;</span>
       </span>`
    : `<span class="av-wrap"><span class="av-ph">&#128100;</span></span>`;

  const pct = Math.round(p.score / MAX * 100);
  const sc  = `<div class="sc-wrap">
    <div class="sc-bar"><div class="sc-fill" style="width:${{pct}}%;background:${{scoreColor(p.score)}}"></div></div>
    <span class="sc-txt">${{p.score}}/${{MAX}}</span>
  </div>`;

  const inds = `<div class="inds">
    ${{ind('📷', p.has_photo,         'Has profile photo')}}
    ${{ind('✏️', p.has_headline,      'Has headline')}}
    ${{ind('🔗', !p.slug_is_generic,  'Custom profile URL')}}
    ${{ind('🤝', p.mutual_level > 0,  MUTUAL_LABELS[Math.min(p.mutual_level,2)])}}
    ${{p.suspicious_name ? `<span class="ind" title="Suspicious name: ${{safeReason}}">⚠️</span>` : ''}}
  </div>`;

  const suspBadge = p.suspicious_name
    ? `<span class="badge badge-susp" title="${{safeReason}}">⚠️ name</span>` : '';

  const noteVal = escapeHtml(notes[p.profile_url] || '');

  return `<tr
    data-idx="${{idx}}"
    data-name="${{escapeHtml(p.name.toLowerCase())}}"
    data-hl="${{escapeHtml((p.headline||'').toLowerCase())}}"
    data-risk="${{p.risk}}"
    data-mutual="${{p.mutual_level}}"
    data-payroll="${{p.payroll_status||''}}">
    <td class="av-cell">${{av}}</td>
    <td><strong>${{safeName}}</strong>${{suspBadge}}</td>
    <td class="hl-cell" title="${{safeHeadline}}">${{safeHeadline || '<em style="color:var(--muted)">—</em>'}}</td>
    <td><span class="badge badge-${{p.risk}}">${{RISK_LABELS[p.risk]}}</span></td>
    <td>${{sc}}</td>
    <td>${{inds}}</td>
    ${{payrollCells(p)}}
    <td><input class="note-input" type="text" placeholder="Add note…" value="${{noteVal}}"></td>
    <td><a class="pl" href="${{safeUrl}}" target="_blank" rel="noopener">View &rarr;</a></td>
  </tr>`;
}}

const tbody = document.getElementById('tbody');
tbody.innerHTML = profiles.map(renderRow).join('');

// Delegated handler — avoids injecting URLs into inline onchange attributes,
// which would break for any profile URL containing quotes or other special
// characters.
tbody.addEventListener('input', (e) => {{
  if (!e.target.classList.contains('note-input')) return;
  const row = e.target.closest('tr');
  const idx = parseInt(row.dataset.idx, 10);
  const url = profiles[idx] && profiles[idx].profile_url;
  if (!url) return;
  const val = e.target.value;
  if (val) notes[url] = val; else delete notes[url];
  saveNotes();
}});

function applyFilters() {{
  const q       = document.getElementById('fSearch').value.toLowerCase();
  const risk    = document.getElementById('fRisk').value;
  const mutual  = document.getElementById('fMutual').value;
  const payroll = hasPayroll ? document.getElementById('fPayroll').value : '';
  let n = 0;
  tbody.querySelectorAll('tr').forEach(row => {{
    const ok = (!q      || row.dataset.name.includes(q) || row.dataset.hl.includes(q))
            && (!risk   || row.dataset.risk    === risk)
            && (!mutual || row.dataset.mutual  === mutual)
            && (!payroll|| row.dataset.payroll === payroll);
    row.style.display = ok ? '' : 'none';
    if (ok) n++;
  }});
  document.getElementById('fCount').textContent = n + ' profile' + (n!==1?'s':'') + ' shown';
  document.getElementById('noRes').style.display = n===0?'block':'none';
}}

['fSearch','fRisk','fMutual'].forEach(id => {{
  const el = document.getElementById(id);
  if (el) el.addEventListener('input', applyFilters);
}});
if (hasPayroll) {{
  const el = document.getElementById('fPayroll');
  if (el) el.addEventListener('change', applyFilters);
}}

function toggleHelp(tab, btn) {{
  const panel = document.getElementById('help-' + tab);
  const wasActive = panel && panel.classList.contains('active');

  // Close all panels and deactivate all tabs first
  document.querySelectorAll('.help-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.help-tab').forEach(t => t.classList.remove('active'));

  // If it wasn't active before, open it; otherwise leave it closed (toggle)
  if (!wasActive && panel) {{
    panel.classList.add('active');
    btn.classList.add('active');
  }}
}}

function exportNotes() {{
  const rows = [['name','profile_url','risk','score','note']];
  let withNotes = 0;
  profiles.forEach(p => {{
    const note = notes[p.profile_url] || '';
    if (note) withNotes++;
    rows.push([p.name, p.profile_url, p.risk, p.score, note]);
  }});
  const csv = rows.map(r => r.map(v => '"'+String(v).replace(/"/g,'""')+'"').join(',')).join('\\n');
  const a = document.createElement('a');
  a.href = 'data:text/csv;charset=utf-8,' + encodeURIComponent('\\uFEFF'+csv);
  a.download = 'headcheck_notes.csv';
  a.click();
  // Informational only — CSV contains every profile so HR can annotate offline.
  console.log('[HeadCheck] Exported ' + profiles.length + ' rows, ' + withNotes + ' with notes.');
}}
</script>
</body>
</html>"""

    with open(out, "w", encoding="utf-8") as f:
        f.write(html)


# ─────────────────────────────────────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf(profiles: list[dict], company: str, has_payroll: bool, out: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    W_PAGE = A4[0]
    W      = W_PAGE - 28*mm
    ts     = datetime.now().strftime("%B %d, %Y")
    total  = len(profiles)
    green  = sum(1 for p in profiles if p["risk"] == "green")
    yellow = sum(1 for p in profiles if p["risk"] == "yellow")
    red    = sum(1 for p in profiles if p["risk"] == "red")

    C_WHITE  = colors.white
    C_BG     = colors.HexColor("#F2F4F8")
    C_SURF   = colors.HexColor("#F8FAFD")
    C_BORDER = colors.HexColor("#D9DEE9")
    C_TEXT   = colors.HexColor("#111827")
    C_MUTED  = colors.HexColor("#4B5563")
    C_BLUE   = colors.HexColor("#1D4ED8")
    C_BLUEDK = colors.HexColor("#1E3A8A")
    C_GREEN  = colors.HexColor("#166534")
    C_AMBER  = colors.HexColor("#92400E")
    C_RED    = colors.HexColor("#991B1B")

    def ps(name, **kw): return ParagraphStyle(name, **kw)

    s_label = ps("lbl",  fontSize=8,  fontName="Helvetica-Bold",    textColor=C_WHITE,  leading=11)
    s_title = ps("ttl",  fontSize=19, fontName="Helvetica-Bold",    textColor=C_TEXT,   leading=23, spaceAfter=3)
    s_co    = ps("co",   fontSize=12, fontName="Helvetica-Bold",    textColor=C_BLUE,   leading=15, spaceAfter=2)
    s_meta  = ps("meta", fontSize=8,  fontName="Helvetica",         textColor=C_MUTED,  leading=12, spaceAfter=2)
    s_h2    = ps("h2",   fontSize=11, fontName="Helvetica-Bold",    textColor=C_TEXT,   leading=14, spaceBefore=10, spaceAfter=5)
    s_small = ps("sm",   fontSize=7,  fontName="Helvetica",         textColor=C_MUTED,  leading=10)
    s_disc  = ps("dc",   fontSize=7,  fontName="Helvetica-Oblique", textColor=C_MUTED,  leading=10)

    def hr(sp=4): return HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=sp, spaceBefore=sp)

    doc = SimpleDocTemplate(out, pagesize=A4,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=14*mm, bottomMargin=16*mm)

    story = []

    # Header band
    hdr = Table([[Paragraph(f"LinkedIn HeadCheck  ·  Security & Workforce Verification  ·  v{VERSION}", s_label)]],
                colWidths=[W])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C_BLUEDK),
        ("TOPPADDING",    (0,0),(-1,-1), 9),
        ("BOTTOMPADDING", (0,0),(-1,-1), 9),
        ("LEFTPADDING",   (0,0),(-1,-1), 12),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("People Audit Report", s_title))
    story.append(Paragraph(company, s_co))
    story.append(Paragraph(f"Generated {ts}  ·  {total} profiles analysed", s_meta))
    story.append(Spacer(1, 3*mm))
    story.append(hr())

    # KPIs
    story.append(Paragraph("Summary", s_h2))

    def kv(val, lbl, color):
        """Return (value_paragraph, label_paragraph) for a KPI cell."""
        return (
            Paragraph(str(val), ps(f"kv{lbl}", fontSize=20, fontName="Helvetica-Bold", textColor=color, leading=22)),
            Paragraph(lbl,      ps(f"kl{lbl}", fontSize=7,  fontName="Helvetica",      textColor=C_MUTED, leading=10)),
        )

    def kpi_table(cells, col_count):
        """Build a two-row KPI table (values on top, labels below) from kv() tuples."""
        values = [c[0] for c in cells]
        labels = [c[1] for c in cells]
        t = Table([values, labels], colWidths=[W/col_count]*col_count)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), C_SURF),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("TOPPADDING",    (0,0),(-1,-1), 10),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
            ("LINEBEFORE",    (1,0),(-1,-1), 0.5, C_BORDER),
            ("LINEBELOW",     (0,0),(-1,0),  0.5, C_BORDER),
            ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
        ]))
        return t

    story.append(kpi_table([
        kv(total,  "Total profiles", C_BLUE),
        kv(green,  "Low risk",       C_GREEN),
        kv(yellow, "Needs review",   C_AMBER),
        kv(red,    "High risk",      C_RED),
    ], col_count=4))
    story.append(Spacer(1, 3*mm))

    if has_payroll:
        exact     = sum(1 for p in profiles if p.get("payroll_status") == "exact")
        fuzzy_n   = sum(1 for p in profiles if p.get("payroll_status") == "fuzzy")
        not_found = sum(1 for p in profiles if p.get("payroll_status") == "not_found")
        story.append(Paragraph("Payroll Cross-Reference", s_h2))
        story.append(kpi_table([
            kv(exact,     "In payroll",     C_GREEN),
            kv(fuzzy_n,   "Similar name",   C_AMBER),
            kv(not_found, "Not in payroll", C_RED),
        ], col_count=3))
        story.append(Spacer(1, 3*mm))

    story.append(hr())

    # Risk explanation
    story.append(Paragraph("Risk Classification", s_h2))
    for lbl, color, desc in [
        ("Low risk (score >= 7)",     C_GREEN, "Complete, active profile. Has photo, custom URL, headline, and/or mutual connections."),
        ("Needs review (score 4-6)",  C_AMBER, "Partial profile. Some signals present but missing key indicators. Manual verification recommended."),
        ("High risk (score <= 3)",    C_RED,   "Ghost or minimal profile, or profile with no photo. Priority for HR investigation."),
    ]:
        story.append(Paragraph(f"<b>{lbl}</b>",
            ps(f"rh{lbl[:3]}", fontSize=9, fontName="Helvetica-Bold", textColor=color, leading=12, spaceBefore=4, spaceAfter=1)))
        story.append(Paragraph(desc, s_small))

    story.append(Spacer(1, 3*mm))
    story.append(hr())

    # Profile tables grouped by risk
    for risk_key, risk_color, section_title in [
        ("red",    C_RED,   "High Risk — Priority Review"),
        ("yellow", C_AMBER, "Needs Review"),
        ("green",  C_GREEN, "Low Risk"),
    ]:
        subset = [p for p in profiles if p["risk"] == risk_key]
        if not subset:
            continue

        story.append(Paragraph(f"{section_title}  ({len(subset)} profiles)", s_h2))

        if has_payroll:
            col_hdrs = ["Name", "Headline", "Profile URL", "Score", "Signals", "Payroll", "Match"]
            col_w    = [W*.18, W*.22, W*.20, W*.08, W*.12, W*.10, W*.10]
        else:
            col_hdrs = ["Name", "Headline", "Profile URL", "Score", "Signals"]
            col_w    = [W*.22, W*.28, W*.28, W*.10, W*.12]

        def th(t): return Paragraph(t, ps(f"th{t[:4]}", fontSize=7, fontName="Helvetica-Bold", textColor=C_MUTED))
        def td(t, bold=False, color=C_TEXT, size=8):
            fn = "Helvetica-Bold" if bold else "Helvetica"
            return Paragraph(str(t), ps(f"td{str(t)[:4]}", fontSize=size, fontName=fn, textColor=color, leading=11))

        rows = [[th(h) for h in col_hdrs]]

        for p in subset:
            sigs, miss = [], []
            if p["has_photo"]:           sigs.append("Photo")
            else:                        miss.append("NO PHOTO")
            if p["has_headline"]:        sigs.append("Title")
            else:                        miss.append("no title")
            if not p["slug_is_generic"]: sigs.append("Custom URL")
            else:                        miss.append("generic URL")
            if p["mutual_level"] == 1:   sigs.append("1 mutual")
            elif p["mutual_level"] >= 2: sigs.append("Multi mutual")
            else:                        miss.append("no mutual")
            if p["suspicious_name"]:     miss.append(f"susp name ({p['suspicious_reason']})")

            sig_txt = (", ".join(sigs) if sigs else "—") + ("\nMissing: " + ", ".join(miss) if miss else "")
            url_short = p["profile_url"].replace("https://www.linkedin.com/in/", "li/in/")

            row = [
                td(p["name"], bold=True),
                td((p["headline"][:65] + "…" if len(p.get("headline","")) > 65 else p.get("headline","")) or "—", color=C_MUTED, size=7),
                Paragraph(f'<link href="{p["profile_url"]}">{url_short}</link>',
                    ps("url", fontSize=7, fontName="Helvetica", textColor=C_BLUE, leading=10)),
                td(f"{p['score']}/{MAX_SCORE}", bold=True, color=risk_color),
                Paragraph(sig_txt, ps("sig", fontSize=7, fontName="Helvetica", textColor=C_TEXT, leading=10)),
            ]
            if has_payroll:
                ps_val = p.get("payroll_status","not_found")
                ps_colors = {"exact":C_GREEN,"fuzzy":C_AMBER,"not_found":C_RED}
                ps_labels = {"exact":"In payroll","fuzzy":"Similar","not_found":"Not found"}
                row.append(td(ps_labels.get(ps_val,"—"), color=ps_colors.get(ps_val,C_MUTED), size=7))
                row.append(td((p.get("payroll_match") or "—")[:22], color=C_MUTED, size=7))

            rows.append(row)

        tbl = Table(rows, colWidths=col_w, repeatRows=1)
        row_styles = [
            ("BACKGROUND",    (0,0),(-1,0),  C_SURF),
            ("LINEBELOW",     (0,0),(-1,0),  1.0, risk_color),
            ("LINEBELOW",     (0,1),(-1,-1), 0.3, C_BORDER),
            ("BOX",           (0,0),(-1,-1), 0.5, C_BORDER),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("LEFTPADDING",   (0,0),(-1,-1), 7),
            ("RIGHTPADDING",  (0,0),(-1,-1), 7),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]
        for i in range(1, len(rows)):
            if i % 2 == 0:
                row_styles.append(("BACKGROUND",(0,i),(-1,i), C_BG))
        tbl.setStyle(TableStyle(row_styles))
        story.append(tbl)
        story.append(Spacer(1, 4*mm))

    story.append(hr())

    # ── NEXT STEPS SECTION ────────────────────────────────────────────────────
    story.append(Paragraph("What to Do After the Audit", s_h2))
    story.append(Paragraph(
        "Once HR has identified suspicious profiles, the following actions are available.",
        ps("ns_intro", fontSize=9, fontName="Helvetica", textColor=C_TEXT, leading=13, spaceAfter=6)
    ))

    # Case 1
    story.append(Paragraph(
        "Former employee who forgot to update their profile",
        ps("ns_h", fontSize=9, fontName="Helvetica-Bold", textColor=C_GREEN, leading=12, spaceBefore=6, spaceAfter=2)
    ))
    story.append(Paragraph(
        "The most common case. Contact the person directly and ask them to edit the Experience section "
        "on their LinkedIn profile and remove or update the position. LinkedIn will automatically "
        "disassociate their profile from the company page. Changes can take up to 30 days to reflect.",
        ps("ns_b", fontSize=8, fontName="Helvetica", textColor=C_TEXT, leading=12, spaceAfter=4)
    ))

    # Case 2
    story.append(Paragraph(
        "Fake or malicious profile",
        ps("ns_h2", fontSize=9, fontName="Helvetica-Bold", textColor=C_RED, leading=12, spaceBefore=4, spaceAfter=2)
    ))
    for step_title, step_body in [
        ("Step 1 — Report to LinkedIn via support form",
         "Page admins cannot remove members directly — only the member can edit their own profile. "
         "Contact LinkedIn providing: the person's full name, a screenshot of the People page showing them, "
         "a link to their profile, and an explanation of why the association is incorrect. "
         "Submit at: linkedin.com/help/linkedin/ask/cp-master "
         "(a confirmed company email address is required)."),
        ("Step 2 — Report the individual profile",
         "On the person's LinkedIn profile, click the More button (···) → Report / Block → "
         "select the appropriate reason (fake profile, incorrect information). "
         "LinkedIn will review and may remove the profile or the company association."),
        ("Step 3 — Escalate if there is no response",
         "If LinkedIn support does not respond within a reasonable time, "
         "contacting @LinkedInHelp on X/Twitter can help speed up the process for clear-cut cases."),
    ]:
        story.append(Paragraph(
            f"<b>{step_title}</b>",
            ps(f"st_{step_title[:4]}", fontSize=8, fontName="Helvetica-Bold", textColor=C_TEXT, leading=12, spaceBefore=3, spaceAfter=1)
        ))
        story.append(Paragraph(
            step_body,
            ps(f"sb_{step_title[:4]}", fontSize=8, fontName="Helvetica", textColor=C_MUTED, leading=12, spaceAfter=3)
        ))

    # What admins cannot do — red-tinted box
    cannot_data = [[Paragraph(
        "<b>What admins cannot do</b>  —  Admins cannot directly remove a member from the company page, "
        "cannot see a full list of associated members in the admin dashboard, and cannot prevent someone "
        "from self-associating with the company. LinkedIn's system relies entirely on self-reported data — "
        "which is exactly why this tool exists.",
        ps("cannot", fontSize=7, fontName="Helvetica", textColor=C_RED, leading=11)
    )]]
    cannot_tbl = Table(cannot_data, colWidths=[W])
    cannot_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#FEF2F2")),
        ("LINEBELOW",     (0,0), (-1,-1), 0, C_BORDER),
        ("LINEBEFORE",    (0,0), (0,-1),  2, C_RED),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
    ]))
    story.append(Spacer(1, 3*mm))
    story.append(cannot_tbl)
    story.append(Spacer(1, 4*mm))

    story.append(hr())
    story.append(Paragraph(
        "Disclaimer: This report was generated automatically from publicly visible LinkedIn profile data. "
        "Risk scores reflect profile completeness only and do not confirm or deny current employment. "
        "All findings must be verified by HR against internal records before any action is taken. "
        "Handle as a confidential internal HR document.",
        s_disc))

    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(C_MUTED)
        canvas.drawString(14*mm, 9*mm, f"LinkedIn HeadCheck  ·  {BRAND_NAME}  ·  {company}  ·  {ts}")
        canvas.drawRightString(W_PAGE - 14*mm, 9*mm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def _evaluate_export_quality(profiles: list[dict], diag: dict) -> list[str]:
    """
    Build a list of human-readable warnings about likely export problems.
    These are shown to the user AFTER extraction, so they can decide whether
    to re-export or trust the results.
    """
    warnings: list[str] = []
    total = len(profiles)
    if not total:
        return warnings

    # Lazy-loaded photos: the user scrolled fast (or not at all).
    not_loaded = sum(1 for p in profiles if p.get("photo_state") == PHOTO_NOT_LOADED)
    if not_loaded / total >= 0.10:
        pct = round(100 * not_loaded / total)
        warnings.append(
            f"{not_loaded} of {total} profiles ({pct}%) have photos that were not "
            "loaded at export time. These are NOT flagged as suspicious, but "
            "the signal is weaker than it could be.\n"
            "           Fix: re-open the People page, scroll slowly to the bottom\n"
            "           so every photo gets a chance to load, then re-export."
        )

    # Mutual-connection coverage: tells us about the exporter, not the profiles.
    with_mutual = sum(1 for p in profiles if p.get("mutual_level", 0) > 0)
    if with_mutual / total < 0.20:
        pct = round(100 * with_mutual / total)
        warnings.append(
            f"Only {with_mutual} of {total} profiles ({pct}%) show mutual "
            "connections. This almost always means the export was done from a\n"
            "           page-admin account (which hides your network) or an account\n"
            "           with limited internal connections. Scores may be systematically\n"
            "           lower than reality.\n"
            "           Fix: have a non-admin employee with broad internal connections\n"
            "           export the page from their own LinkedIn session."
        )

    # No profiles of a given base: might indicate LinkedIn changed its DOM.
    if diag.get("anchors_found", 0) > 0 and total == 0:
        warnings.append(
            "Profile card anchors were found, but no profiles could be parsed.\n"
            "           LinkedIn may have changed its HTML structure. Run with --debug\n"
            "           to see the extraction diagnostics."
        )

    return warnings


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

# Progress stages the callback receives as its first argument.
STAGE_EXTRACT   = "extract"
STAGE_PAYROLL   = "payroll"
STAGE_REPORTS   = "reports"
STAGE_XLSX      = "xlsx"
STAGE_SUSPECTS  = "suspects"
STAGE_SNAPSHOT  = "snapshot"
STAGE_DONE      = "done"


def run_headcheck(
    html_path: str,
    company: str = "Company",
    payroll_path: str | None = None,
    lang: str = "en",
    out_dir: str = ".",
    progress: callable = None,
) -> dict:
    """
    Run the full HeadCheck pipeline and return a structured result.

    This function is the programmatic entry point for both the CLI and the
    TUI. It does NOT print anything — callers get all information via the
    returned dict, and may optionally pass a `progress` callback that will
    be invoked at each stage boundary with (stage_name, details_dict).

    Parameters
    ----------
    html_path : str
        Path to the LinkedIn People page HTML snapshot.
    company : str
        Company name used in report headers and output filenames.
    payroll_path : str | None
        Optional path to a .csv/.xlsx/.xls payroll file for cross-reference.
    lang : str
        Language code of the LinkedIn interface used during export. Must be
        one of MUTUAL_PATTERNS.keys() — defaults to "en".
    out_dir : str
        Directory where the three output files will be written. Created if
        it doesn't exist.
    progress : callable | None
        Optional callback(stage: str, info: dict) invoked between stages.
        See STAGE_* constants.

    Returns
    -------
    dict with keys:
        config:    the input parameters (echoed back for UI convenience)
        profiles:  full list of extracted profile dicts, scored
        diagnostics: the raw extractor diagnostics (counters, skip reasons)
        warnings:  list of export-quality warning strings
        payroll:   {loaded: int, detected_col: int | None, has_payroll: bool}
        outputs:   {html: path, pdf: path, suspects_csv: path}
        stats:     {total, red, yellow, green, suspects_exported}
    """
    def _emit(stage: str, **info):
        if progress is not None:
            progress(stage, info)

    os.makedirs(out_dir, exist_ok=True)
    date_slug = datetime.now().strftime("%Y-%m-%d")
    co_slug   = re.sub(r"[^a-z0-9]+", "_", company.lower()).strip("_") or "company"
    base      = f"headcheck_{co_slug}_{date_slug}"

    # ── Stage 1: extract ──
    profiles, diag = extract_profiles(html_path, lang)
    warnings_list = _evaluate_export_quality(profiles, diag)
    _emit(STAGE_EXTRACT, count=len(profiles), diagnostics=diag, warnings=warnings_list)

    # ── Stage 2: payroll (optional) ──
    payroll_info = {"loaded": 0, "detected_col": None, "has_payroll": False}
    if payroll_path:
        names, col_idx = load_payroll_detailed(payroll_path)
        payroll_info["loaded"] = len(names)
        payroll_info["detected_col"] = col_idx
        if names:
            profiles = cross_reference(profiles, names)
            payroll_info["has_payroll"] = True
        _emit(STAGE_PAYROLL, **payroll_info)
    else:
        _emit(STAGE_PAYROLL, **payroll_info)

    has_payroll = payroll_info["has_payroll"]

    # ── Stage 3: reports ──
    html_out = os.path.join(out_dir, f"{base}.html")
    pdf_out  = os.path.join(out_dir, f"{base}.pdf")
    generate_html(profiles, company, has_payroll, html_out)
    generate_pdf(profiles, company, has_payroll, pdf_out)
    _emit(STAGE_REPORTS, html=html_out, pdf=pdf_out)

    # ── Stage 4: Excel workbook for HR ──
    xlsx_out = os.path.join(out_dir, f"{base}.xlsx")
    xlsx_rows = generate_xlsx(profiles, company, has_payroll, xlsx_out)
    _emit(STAGE_XLSX, path=xlsx_out, rows=xlsx_rows)

    # ── Stage 5: suspects CSV ──
    csv_out = os.path.join(out_dir, f"{base}_suspects.csv")
    suspects_count = export_suspects_csv(profiles, has_payroll, csv_out)
    _emit(STAGE_SUSPECTS, path=csv_out, count=suspects_count)

    # ── Stats (needed by the snapshot) ──
    stats = {
        "total":  len(profiles),
        "red":    sum(1 for p in profiles if p["risk"] == "red"),
        "yellow": sum(1 for p in profiles if p["risk"] == "yellow"),
        "green":  sum(1 for p in profiles if p["risk"] == "green"),
        "suspects_exported": suspects_count,
    }

    # ── Stage 6: snapshot JSON for later diffing ──
    snap_out = os.path.join(out_dir, f"{base}.json")
    export_snapshot_json(profiles, company, has_payroll, stats, snap_out)
    _emit(STAGE_SNAPSHOT, path=snap_out, count=len(profiles))

    result = {
        "config": {
            "html_path": html_path,
            "company": company,
            "payroll_path": payroll_path,
            "lang": lang,
            "out_dir": out_dir,
        },
        "profiles": profiles,
        "diagnostics": diag,
        "warnings": warnings_list,
        "payroll": payroll_info,
        "outputs": {"html": html_out, "pdf": pdf_out,
                    "xlsx": xlsx_out, "suspects_csv": csv_out,
                    "snapshot": snap_out},
        "stats": stats,
    }
    _emit(STAGE_DONE, **stats)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _cli_progress_printer(debug: bool = False):
    """
    Build a progress callback that reproduces the classic CLI output style.
    Kept as a module-level factory so the TUI can construct its own callback
    independently without any CLI-specific strings leaking.
    """
    def printer(stage: str, info: dict):
        if stage == STAGE_EXTRACT:
            count = info["count"]
            print("[ 1/6 ] Extracting profiles…")
            print(f"        {count} unique profiles found")
            if count == 0:
                print(
                    "        ⚠  No profiles were extracted. The HTML snapshot may be\n"
                    "           incomplete (did you scroll to the bottom before copying?)\n"
                    "           or LinkedIn may have changed its DOM structure.\n"
                )
            if debug:
                print("   [debug] extraction diagnostics:", info["diagnostics"])
            if info["warnings"]:
                print("        ⚠  Export-quality warnings:")
                for w in info["warnings"]:
                    print(f"           • {w}\n")

        elif stage == STAGE_PAYROLL:
            if info["has_payroll"]:
                print("[ 2/6 ] Loading payroll…")
                print(f"   Auto-detected name column index: {info['detected_col']}")
                print(f"        {info['loaded']} employees loaded")
            elif info["loaded"] == 0 and info["detected_col"] is None:
                print("[ 2/6 ] No payroll file — skipping cross-reference")
            else:
                # payroll path given but no names found
                print("[ 2/6 ] Loading payroll…")
                print("        ⚠  Payroll file parsed but no names were detected — skipping cross-reference")

        elif stage == STAGE_REPORTS:
            print("[ 3/6 ] Generating reports…")
            print(f"  ✔ HTML report  →  {info['html']}")
            print(f"  ✔ PDF report   →  {info['pdf']}")

        elif stage == STAGE_XLSX:
            print("[ 4/6 ] Generating Excel workbook for HR…")
            print(f"  ✔ Excel        →  {info['path']}  ({info['rows']} profiles)")

        elif stage == STAGE_SUSPECTS:
            print("[ 5/6 ] Exporting suspects list…")
            print(f"  ✔ Suspects CSV →  {info['path']}  ({info['count']} profiles)")

        elif stage == STAGE_SNAPSHOT:
            print("[ 6/6 ] Saving snapshot for future diffs…")
            print(f"  ✔ JSON         →  {info['path']}  ({info['count']} profiles)")

    return printer


# ─────────────────────────────────────────────────────────────────────────────
# INTERACTIVE MODE (questionary + rich)
# ─────────────────────────────────────────────────────────────────────────────
#
# The interactive mode is activated when `headcheck` is invoked with no
# arguments. It wraps run_headcheck() with a questionary-driven wizard and a
# rich-powered progress display.
#
# Both libraries are imported lazily so users who only rely on the classic
# CLI don't need to install them. If questionary or rich is missing, we fall
# back to a minimal plain-stdin wizard that works with nothing but the
# standard library.


def _try_import_interactive():
    """
    Attempt to import questionary and rich. Returns (questionary, rich_console
    Console class) on success, (None, None) if either is missing.
    """
    try:
        import questionary
        from rich.console import Console
        return questionary, Console
    except ImportError:
        return None, None


def _plain_wizard(console_print) -> dict:
    """
    Zero-dependency fallback wizard. Used when questionary / rich aren't
    installed. Same questions, much uglier input loop.
    """
    console_print("\n  LinkedIn HeadCheck — interactive mode (plain fallback)")
    console_print("  Tip: install `questionary` and `rich` for a nicer experience.\n")

    while True:
        html = input("  HTML snapshot path: ").strip()
        if html and os.path.isfile(html):
            break
        print("  ✗ File not found. Try again.")

    company = input("  Company name [Company]: ").strip() or "Company"

    payroll = input("  Payroll file (optional, blank to skip): ").strip() or None
    if payroll and not os.path.isfile(payroll):
        print(f"  ⚠  {payroll!r} not found — continuing without payroll.")
        payroll = None

    langs = list(MUTUAL_PATTERNS.keys())
    lang = input(f"  LinkedIn language {langs} [en]: ").strip() or "en"
    if lang not in langs:
        lang = "en"

    out = input("  Output directory [.]: ").strip() or "."

    return dict(html_path=html, company=company, payroll_path=payroll,
                lang=lang, out_dir=out)


def _questionary_wizard(questionary) -> dict | None:
    """
    Friendly wizard powered by questionary. Returns None if the user aborts
    (Ctrl+C or ESC on any prompt), else the configuration dict.
    """
    def _file_exists(v: str) -> bool | str:
        if not v:
            return "This field is required"
        if not os.path.isfile(v):
            return f"File not found: {v}"
        return True

    def _file_or_empty(v: str) -> bool | str:
        if not v:
            return True
        if not os.path.isfile(v):
            return f"File not found: {v}"
        return True

    answers = questionary.form(
        html_path=questionary.path(
            "HTML snapshot path:",
            validate=_file_exists,
        ),
        company=questionary.text(
            "Company name:",
            default="Company",
            instruction="(used in report headers and filenames)",
        ),
        payroll_path=questionary.path(
            "Payroll file (optional, leave blank to skip):",
            default="",
            validate=_file_or_empty,
        ),
        lang=questionary.select(
            "LinkedIn interface language used during export:",
            choices=list(MUTUAL_PATTERNS.keys()),
            default="en",
        ),
        out_dir=questionary.path(
            "Output directory:",
            default=".",
        ),
    ).ask()

    if answers is None:                          # user hit Ctrl+C / ESC
        return None
    # questionary.path() returns "" for skipped optional fields — normalise.
    answers["payroll_path"] = answers["payroll_path"] or None
    return answers


def _rich_progress_callback(console):
    """
    Build a progress callback that renders each stage as a rich status line.
    Closed over `console` so both header and per-stage output share the same
    terminal handler (important for colour / width detection).
    """
    from rich.panel import Panel
    from rich.table import Table

    def cb(stage: str, info: dict):
        if stage == STAGE_EXTRACT:
            console.print(f"[bold cyan]⬢ Extracting profiles[/bold cyan] "
                          f"· {info['count']} unique")
            if info["count"] == 0:
                console.print(
                    "  [yellow]⚠  No profiles were extracted. The HTML snapshot may "
                    "be incomplete (scroll to the bottom before exporting) or "
                    "LinkedIn may have changed its DOM structure.[/yellow]"
                )
            for w in info["warnings"]:
                console.print(Panel(w, title="Export quality warning",
                                    title_align="left", border_style="yellow"))

        elif stage == STAGE_PAYROLL:
            if info["has_payroll"]:
                console.print(f"[bold cyan]⬢ Payroll[/bold cyan] · "
                              f"{info['loaded']} employees "
                              f"(name column #{info['detected_col']})")
            elif info["loaded"] == 0 and info["detected_col"] is None:
                console.print("[dim]⬢ Payroll skipped (no file provided)[/dim]")
            else:
                console.print("[yellow]⬢ Payroll file had no detectable names — "
                              "cross-reference skipped[/yellow]")

        elif stage == STAGE_REPORTS:
            console.print(f"[bold cyan]⬢ Reports[/bold cyan]")
            console.print(f"  [green]✔[/green] HTML  → {info['html']}")
            console.print(f"  [green]✔[/green] PDF   → {info['pdf']}")

        elif stage == STAGE_XLSX:
            console.print(f"[bold cyan]⬢ Excel workbook for HR[/bold cyan] · "
                          f"{info['rows']} profiles")
            console.print(f"  [green]✔[/green] XLSX  → {info['path']}")

        elif stage == STAGE_SUSPECTS:
            console.print(f"[bold cyan]⬢ Suspects list[/bold cyan] · "
                          f"{info['count']} profiles")
            console.print(f"  [green]✔[/green] CSV   → {info['path']}")

        elif stage == STAGE_SNAPSHOT:
            console.print(f"[bold cyan]⬢ Snapshot[/bold cyan] · "
                          f"saved for future diffs")
            console.print(f"  [green]✔[/green] JSON  → {info['path']}")

        elif stage == STAGE_DONE:
            table = Table(show_header=False, box=None, padding=(0, 2))
            table.add_row("[bold red]🔴 High risk[/bold red]",    str(info["red"]))
            table.add_row("[bold yellow]🟡 Needs review[/bold yellow]", str(info["yellow"]))
            table.add_row("[bold green]🟢 Low risk[/bold green]",  str(info["green"]))
            console.print()
            console.print(Panel(table, title="Results",
                                title_align="left", border_style="cyan"))

    return cb


def _run_interactive() -> int:
    """
    Interactive entry point. Returns an exit code.
    """
    questionary, Console = _try_import_interactive()

    if Console is not None:
        console = Console()
        _print = console.print
        # Banner
        console.print()
        console.print(f"[bold cyan]LinkedIn HeadCheck[/bold cyan] "
                      f"[dim]v{VERSION}[/dim] · {BRAND_NAME}")
        console.print(f"[dim]{'─' * 46}[/dim]\n")
    else:
        console = None
        _print = print
        _print(f"\n  LinkedIn HeadCheck  v{VERSION}  ·  {BRAND_NAME}")
        _print(f"  {'─' * 46}\n")

    # Collect answers
    if questionary is not None:
        answers = _questionary_wizard(questionary)
        if answers is None:
            _print("\n[dim]Cancelled.[/dim]" if console else "\n  Cancelled.")
            return 1
    else:
        # One-time hint: the plain wizard works but is uglier.
        _print("  (Install `questionary` and `rich` for a better experience:")
        _print("   pip install questionary rich)\n")
        answers = _plain_wizard(_print)

    # Run the pipeline with the right progress renderer
    if console is not None:
        progress = _rich_progress_callback(console)
    else:
        progress = _cli_progress_printer()

    console and console.print()
    try:
        run_headcheck(progress=progress, **answers)
    except Exception as e:                       # noqa: BLE001 — user-facing
        if console is not None:
            console.print(f"\n[bold red]Error:[/bold red] {e}")
        else:
            print(f"\n  Error: {e}")
        return 2
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# DIFF SUBCOMMAND
# ─────────────────────────────────────────────────────────────────────────────

def _format_diff_plain(diff: dict) -> str:
    """Render a snapshot diff as plain text — no colour, works anywhere."""
    meta = diff["meta"]
    lines = []
    lines.append("")
    lines.append("  HeadCheck — Snapshot Diff")
    lines.append("  " + "─" * 46)
    lines.append(f"  Old: {meta['old_path']}  ({meta.get('old_generated_at') or 'unknown date'})")
    lines.append(f"  New: {meta['new_path']}  ({meta.get('new_generated_at') or 'unknown date'})")
    lines.append("")

    def _section(title: str, items: list, formatter):
        if not items:
            return
        lines.append(f"  {title}  ({len(items)})")
        lines.append("  " + "─" * 46)
        for p in items:
            lines.append("    " + formatter(p))
        lines.append("")

    _section("✚ Appeared (new since last audit)",
             diff["appeared"],
             lambda p: f"{p.get('name',''):30s}  risk={p.get('risk','?'):6s}  {p.get('profile_url','')}")
    _section("✖ Disappeared (no longer in company)",
             diff["disappeared"],
             lambda p: f"{p.get('name',''):30s}  risk={p.get('risk','?'):6s}  {p.get('profile_url','')}")
    _section("⬆ Risk worsened",
             diff["risk_up"],
             lambda p: f"{p['name']:30s}  {p['old_risk']} → {p['new_risk']}   score {p['old_score']} → {p['new_score']}")
    _section("⬇ Risk improved",
             diff["risk_down"],
             lambda p: f"{p['name']:30s}  {p['old_risk']} → {p['new_risk']}   score {p['old_score']} → {p['new_score']}")
    _section("~ Score drift (same risk)",
             diff["score_changed"],
             lambda p: f"{p['name']:30s}  score {p['old_score']} → {p['new_score']}  ({p['new_risk']})")

    total_changes = sum(len(diff[k]) for k in
                        ("appeared", "disappeared", "risk_up", "risk_down", "score_changed"))
    if total_changes == 0:
        lines.append("  Nothing changed between the two snapshots.")
        lines.append("")
    else:
        lines.append(f"  Summary: {total_changes} changes · "
                     f"{len(diff['unchanged'])} profiles unchanged")
        lines.append("")
    return "\n".join(lines)


def _format_diff_rich(diff: dict):
    """Render a diff using rich — colour, panels, nicer typography."""
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel

    console = Console()
    meta = diff["meta"]

    console.print()
    console.print("[bold cyan]HeadCheck — Snapshot Diff[/bold cyan]")
    console.print(f"[dim]{'─' * 46}[/dim]")
    console.print(f"[dim]Old:[/dim] {meta['old_path']}  [dim]({meta.get('old_generated_at') or '—'})[/dim]")
    console.print(f"[dim]New:[/dim] {meta['new_path']}  [dim]({meta.get('new_generated_at') or '—'})[/dim]")
    console.print()

    def _section(title: str, items: list, cols: list, extractor, style: str):
        if not items:
            return
        t = Table(title=f"{title}  ({len(items)})", title_style=style, title_justify="left",
                  show_header=True, header_style="bold", box=None, padding=(0, 1))
        for col in cols:
            t.add_column(col)
        for p in items:
            t.add_row(*extractor(p))
        console.print(t)
        console.print()

    _section(
        "✚ Appeared (new since last audit)",
        diff["appeared"],
        ["Name", "Risk", "URL"],
        lambda p: (p.get("name", ""), p.get("risk", "?"), p.get("profile_url", "")),
        "bold green",
    )
    _section(
        "✖ Disappeared (no longer listed)",
        diff["disappeared"],
        ["Name", "Was", "URL"],
        lambda p: (p.get("name", ""), p.get("risk", "?"), p.get("profile_url", "")),
        "bold magenta",
    )
    _section(
        "⬆ Risk worsened",
        diff["risk_up"],
        ["Name", "Risk", "Score"],
        lambda p: (p["name"], f"{p['old_risk']} → {p['new_risk']}",
                   f"{p['old_score']} → {p['new_score']}"),
        "bold red",
    )
    _section(
        "⬇ Risk improved",
        diff["risk_down"],
        ["Name", "Risk", "Score"],
        lambda p: (p["name"], f"{p['old_risk']} → {p['new_risk']}",
                   f"{p['old_score']} → {p['new_score']}"),
        "bold green",
    )
    _section(
        "~ Score drift (same risk)",
        diff["score_changed"],
        ["Name", "Score", "Risk"],
        lambda p: (p["name"], f"{p['old_score']} → {p['new_score']}", p["new_risk"]),
        "bold yellow",
    )

    total_changes = sum(len(diff[k]) for k in
                        ("appeared", "disappeared", "risk_up", "risk_down", "score_changed"))
    unchanged = len(diff["unchanged"])
    if total_changes == 0:
        console.print(Panel("Nothing changed between the two snapshots.",
                            border_style="dim"))
    else:
        console.print(Panel(f"[bold]{total_changes}[/bold] changes · "
                            f"[dim]{unchanged} profiles unchanged[/dim]",
                            title="Summary", title_align="left", border_style="cyan"))


def _run_diff(argv: list[str]) -> int:
    """
    Entry point for `headcheck diff old.json new.json [--csv path]`.
    Returns a shell-friendly exit code:
        0 — snapshots compared, no changes
        1 — snapshots compared, at least one change detected
        2 — bad usage or I/O error
    """
    parser = argparse.ArgumentParser(
        prog="headcheck diff",
        description="Compare two HeadCheck JSON snapshots and report what changed.",
    )
    parser.add_argument("old", help="Path to the older snapshot (.json)")
    parser.add_argument("new", help="Path to the newer snapshot (.json)")
    parser.add_argument("--csv", help="Optional path to also write a CSV export of the diff")
    parser.add_argument("--plain", action="store_true",
                        help="Plain text output (no colour, no rich)")
    args = parser.parse_args(argv)

    try:
        diff = diff_snapshots(args.old, args.new)
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 2
    except json.JSONDecodeError as e:
        print(f"Error: invalid JSON in snapshot file — {e}", file=sys.stderr)
        return 2

    # Try rich unless the user asked for plain output or rich is missing.
    used_rich = False
    if not args.plain:
        try:
            _format_diff_rich(diff)
            used_rich = True
        except ImportError:
            pass
    if not used_rich:
        print(_format_diff_plain(diff))

    if args.csv:
        n = export_diff_csv(diff, args.csv)
        print(f"\n  ✔ Diff CSV  →  {args.csv}  ({n} rows)")

    has_changes = any(diff[k] for k in
                      ("appeared", "disappeared", "risk_up", "risk_down", "score_changed"))
    return 1 if has_changes else 0


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Subcommand dispatch: `headcheck diff ...` runs the snapshot comparator.
    # The usual flag-based CLI and the interactive wizard are the default paths.
    if len(sys.argv) >= 2 and sys.argv[1] == "diff":
        sys.exit(_run_diff(sys.argv[2:]))

    # When the user invokes `headcheck` with no arguments, drop into the
    # interactive wizard. The CLI with flags stays untouched for scripting
    # and for users who prefer it.
    if len(sys.argv) == 1:
        sys.exit(_run_interactive())

    parser = argparse.ArgumentParser(
        prog="headcheck",
        description="LinkedIn HeadCheck — Workforce Verification Tool by Not Nulled Labs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Run with no arguments for an interactive wizard, or use flags below for
scripting:

  python headcheck.py                                              (interactive)
  python headcheck.py --html people.html --company "Acme Corp"
  python headcheck.py --html people.html --company "Acme Corp" --payroll staff.xlsx
  python headcheck.py --html people.html --company "Acme Corp" --payroll nomina.csv --lang es
  python headcheck.py --html people.html --company "Acme Corp" --out ./reports

Compare two audits with the diff subcommand:

  python headcheck.py diff old.json new.json
  python headcheck.py diff old.json new.json --csv changes.csv

Supported payroll formats: .csv  .xlsx  .xls
Supported languages (--lang): en  es  fr  de  pt  it

{BRAND_NAME}  ·  {BRAND_URL}
{REPO_URL}
        """
    )
    parser.add_argument("--html",    required=True,   help="Path to saved LinkedIn People page HTML")
    parser.add_argument("--company", default="Company", help="Company name for report headers")
    parser.add_argument("--payroll", default=None,    help="Optional payroll file (.csv, .xlsx, .xls)")
    parser.add_argument("--lang",    default="en",    choices=list(MUTUAL_PATTERNS.keys()),
                        help="Language of the LinkedIn interface used during export (default: en)")
    parser.add_argument("--out",     default=".",     help="Output directory (default: current directory)")
    parser.add_argument("--debug",   action="store_true",
                        help="Print extraction diagnostics to help pinpoint parser issues")
    args = parser.parse_args()

    # Header block — keeps the current look of the CLI intact.
    print(f"\n  LinkedIn HeadCheck  v{VERSION}  ·  {BRAND_NAME}")
    print(f"  {'─'*46}")
    print(f"  Company  : {args.company}")
    print(f"  HTML     : {args.html}")
    print(f"  Language : {args.lang}")
    if args.payroll:
        print(f"  Payroll  : {args.payroll}")
    print(f"  Output   : {args.out}\n")

    result = run_headcheck(
        html_path=args.html,
        company=args.company,
        payroll_path=args.payroll,
        lang=args.lang,
        out_dir=args.out,
        progress=_cli_progress_printer(debug=args.debug),
    )

    s = result["stats"]
    print(f"\n  Done.  🔴 {s['red']} high-risk  ·  🟡 {s['yellow']} review  ·  🟢 {s['green']} low-risk")
    print(f"  Reports saved to: {args.out}\n")


if __name__ == "__main__":
    main()
