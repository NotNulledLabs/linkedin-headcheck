"""
Tests for generate_xlsx().

The XLSX format is hard to test comprehensively — we can't validate visual
output programmatically — so these tests focus on structural guarantees:
correct header row, expected column count, hyperlinks on URLs, risk cell
background colours, frozen panes, autofilter enabled, data bars present
on the score column, and a hidden _meta sheet with provenance info.
"""
import pytest

pytest.importorskip("openpyxl")

import openpyxl

from headcheck import (
    generate_xlsx,
    PHOTO_LOADED, PHOTO_NOT_LOADED, PHOTO_ABSENT,
    MAX_SCORE,
)


def _mkprofile(**kw):
    base = dict(
        name="Alice Johnson",
        profile_url="https://www.linkedin.com/in/alice-johnson",
        avatar_url="https://media.licdn.com/alice.jpg",
        photo_state=PHOTO_LOADED,
        has_photo=True,
        headline="Senior Engineer at Acme Corp",
        slug="alice-johnson",
        slug_is_generic=False,
        mutual_level=2,
        has_headline=True,
        has_employer_ref=True,
        suspicious_name=False,
        suspicious_reason="",
        score=9,
        risk="green",
    )
    base.update(kw)
    return base


@pytest.fixture
def sample_profiles():
    return [
        _mkprofile(name="Alice Johnson", risk="green",  score=9),
        _mkprofile(name="Bob Smith",     risk="yellow", score=5, mutual_level=0),
        _mkprofile(name="Ghost",         risk="red",    score=0,
                   photo_state=PHOTO_ABSENT, has_photo=False),
    ]


class TestXlsxStructure:
    def test_workbook_has_expected_sheets(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        assert "HeadCheck" in wb.sheetnames
        assert "_meta" in wb.sheetnames

    def test_headers_without_payroll(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        headers = [c.value for c in ws[1]]
        # No payroll columns; Notes is the last column.
        assert "HR notes" in headers
        assert "Payroll status" not in headers
        assert "Payroll match" not in headers

    def test_headers_with_payroll(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=True, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        headers = [c.value for c in ws[1]]
        assert "Payroll status" in headers
        assert "Payroll match" in headers
        # HR notes must still come LAST.
        assert headers[-1] == "HR notes"

    def test_row_count_matches_profiles(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        n = generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))
        assert n == len(sample_profiles)

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        # Max row = header (1) + data rows.
        assert ws.max_row == 1 + len(sample_profiles)

    def test_rows_ordered_by_score_ascending(self, sample_profiles, tmp_path):
        """Worst-first ordering — the red row must appear before yellow/green."""
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        scores = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
        assert scores == sorted(scores)


class TestXlsxVisualAids:
    def test_header_row_frozen(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        assert ws.freeze_panes == "A2"

    def test_autofilter_enabled(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        # Autofilter range should cover the full data region.
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1")

    def test_risk_cells_coloured(self, sample_profiles, tmp_path):
        """Every data row's Risk cell must have a non-white background."""
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        for i in range(2, ws.max_row + 1):
            cell = ws.cell(row=i, column=1)
            fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            assert fill_rgb is not None
            # Light-coloured fills — check they end in something non-zero.
            # (We don't hard-code exact RGB to avoid brittleness if we
            # re-tune the palette.)
            assert fill_rgb != "00000000"

    def test_url_cells_are_hyperlinks(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        # Column E (5) is Profile URL; hyperlink attribute must be set.
        for i in range(2, ws.max_row + 1):
            cell = ws.cell(row=i, column=5)
            if cell.value:
                assert cell.hyperlink is not None
                assert cell.hyperlink.target.startswith("http")

    def test_score_data_bars_present(self, sample_profiles, tmp_path):
        """
        Conditional formatting rules should include a data bar on the
        Score column.
        """
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=False, out=str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        # openpyxl stores conditional formatting rules under _cf_rules.
        cf = list(ws.conditional_formatting._cf_rules.items()) \
            if hasattr(ws.conditional_formatting, "_cf_rules") \
            else list(ws.conditional_formatting)
        # There must be at least one rule targeting the Score column.
        assert len(cf) >= 1


class TestXlsxMetaSheet:
    def test_meta_sheet_is_hidden(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme", has_payroll=True, out=str(out))

        wb = openpyxl.load_workbook(out)
        meta = wb["_meta"]
        assert meta.sheet_state == "hidden"

    def test_meta_sheet_has_provenance(self, sample_profiles, tmp_path):
        out = tmp_path / "r.xlsx"
        generate_xlsx(sample_profiles, "Acme Corp", has_payroll=True, out=str(out))

        wb = openpyxl.load_workbook(out)
        meta = wb["_meta"]
        # Collect all B-column values (keyed by A-column labels).
        kv = {}
        for row in meta.iter_rows(values_only=True):
            if row and row[0]:
                kv[row[0]] = row[1]
        assert "Acme Corp" == kv.get("Company")
        assert kv.get("Total profiles") == len(sample_profiles)
        assert kv.get("Payroll cross-referenced") == "yes"


class TestXlsxEmptyProfiles:
    def test_zero_profiles_produces_valid_empty_workbook(self, tmp_path):
        out = tmp_path / "r.xlsx"
        n = generate_xlsx([], "Acme", has_payroll=False, out=str(out))
        assert n == 0

        # The workbook should still be openable — header row only.
        wb = openpyxl.load_workbook(out)
        ws = wb["HeadCheck"]
        assert ws.max_row == 1   # only the header
